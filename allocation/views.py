from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Count, Q, Avg, Sum, Case, When, IntegerField, F
from django.db import IntegrityError
from django.core.files.storage import default_storage
from django.conf import settings
from decimal import Decimal, InvalidOperation
from urllib.parse import urlencode
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.views.decorators.csrf import ensure_csrf_cookie
from django.contrib.staticfiles import finders
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle # type: ignore
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from .forms import (
    CEForm, StudentForm, StudentSearchForm, AttendanceForm, ExcelUploadForm,
    StudentUploadForm, MarksheetUploadForm, ResultLookupForm, SubjectForm,
    NoticeForm, FacultyMarksEntryForm
)
from .models import (
    CESeating, Student, Faculty, Subject, Semester, Branch,
    Attendance, StudentSubject, ResultSheet, ResultEntry,
    Notice, PushSubscription, NoticeAttachment, StudentMark,
    MentorActionLog, MarksFreezeRule, MarksAuditTrail, MentorAssignment
)
import pandas as pd
import io
import csv
import json
from datetime import datetime, timedelta
import calendar
import re
import hashlib
import sys

try:
    from pywebpush import webpush, WebPushException
except Exception:
    webpush = None
    WebPushException = Exception


def _clean_cell(value):
    if value is None:
        return ''
    try:
        if pd.isna(value):
            return ''
    except Exception:
        pass
    return str(value).strip()


def _normalize_text(value):
    text = _clean_cell(value).lower()
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def _build_attendance_summary_pdf_response(institute_line1, institute_line2, exam_title, semester, table_rows):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="ce_seating_summary.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 40

    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, y, institute_line1)
    y -= 18
    p.drawCentredString(width / 2, y, institute_line2)

    y -= 22
    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width / 2, y, exam_title or "Mid Semester Exam")
    y -= 16
    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width / 2, y, "Seating Arrangement")

    y -= 18
    p.setFont("Helvetica", 9)
    p.drawString(40, y, f"Semester: {semester or '-'}")
    p.drawRightString(width - 40, y, "Date: ________________")

    y -= 18
    pdf_table_data = [['Branch', 'Block No', 'Room No', 'Seat Nos', 'No. of Students']]
    branch_start_row = 1
    for row in table_rows:
        if not row.get('is_total'):
            seat_from = _clean_cell(row.get('seat_from'))
            seat_to = _clean_cell(row.get('seat_to'))
            if seat_from and seat_to and seat_from != seat_to:
                seat_text = f"{seat_from} - {seat_to}"
            else:
                seat_text = seat_from or seat_to
            pdf_table_data.append([
                row.get('branch', ''),
                row.get('block_no', ''),
                row.get('room_no', ''),
                seat_text,
                row.get('count', ''),
            ])
        else:
            pdf_table_data.append(['', '', '', 'Total', row.get('count', '')])

    branch_end_row = len(pdf_table_data) - 2
    table = Table(pdf_table_data, colWidths=[1.1 * inch, 1.2 * inch, 1.2 * inch, 2.7 * inch, 1.4 * inch])
    spans = []
    if branch_end_row >= branch_start_row:
        spans.append(('SPAN', (0, branch_start_row), (0, branch_end_row)))

    table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 10),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('FONT', (0, -1), (-1, -1), 'Helvetica-Bold', 9),
        ('GRID', (0, 0), (-1, -1), 0.8, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ] + spans))

    table_width = 7.6 * inch
    table.wrapOn(p, table_width, height)
    table.drawOn(p, (width - table_width) / 2, max(100, y - (20 * len(pdf_table_data))))

    p.setFont("Helvetica", 9)
    p.drawCentredString(width / 2, 30, "Prepared by Exam Section")
    p.save()
    return response


def _roman_to_int(roman):
    if not roman:
        return None
    roman = str(roman).upper().strip()
    values = {'I': 1, 'V': 5, 'X': 10}
    total = 0
    prev = 0
    for ch in reversed(roman):
        if ch not in values:
            return None
        val = values[ch]
        if val < prev:
            total -= val
        else:
            total += val
            prev = val
    return total if 1 <= total <= 20 else None


def _find_semester_and_session(raw_df):
    scan_rows = min(len(raw_df.index), 20)
    semester_no = None
    exam_session = None

    for i in range(scan_rows):
        row_text = ' '.join(_clean_cell(x) for x in raw_df.iloc[i].tolist() if _clean_cell(x))
        if not row_text:
            continue

        sem_match = re.search(r'semester\s*[:\-]?\s*([ivx]+|\d+)', row_text, flags=re.IGNORECASE)
        if sem_match and semester_no is None:
            sem_token = sem_match.group(1)
            if sem_token.isdigit():
                semester_no = int(sem_token)
            else:
                semester_no = _roman_to_int(sem_token)

        session_match = re.search(
            r'(JANUARY|FEBRUARY|MARCH|APRIL|MAY|JUNE|JULY|AUGUST|SEPTEMBER|OCTOBER|NOVEMBER|DECEMBER)\s+\d{4}',
            row_text,
            flags=re.IGNORECASE,
        )
        if session_match and exam_session is None:
            exam_session = session_match.group(0).upper()

    return semester_no, exam_session


def _looks_like_enrollment(value):
    token = re.sub(r'[^A-Za-z0-9]', '', _clean_cell(value)).upper()
    return len(token) >= 8 and any(ch.isalpha() for ch in token) and any(ch.isdigit() for ch in token)


def _parse_number(value, as_int=False):
    text = _clean_cell(value)
    if not text:
        return None
    try:
        number = float(text)
    except Exception:
        return None
    if as_int:
        return int(round(number))
    return number


def _extract_consolidated_marks_rows(raw_df):
    if raw_df is None or raw_df.empty:
        return None, 'Uploaded sheet is empty.'

    semester_no, exam_session = _find_semester_and_session(raw_df)

    header_row = None
    for i in range(min(len(raw_df.index), 60)):
        row_tokens = [_normalize_text(x) for x in raw_df.iloc[i].tolist()]
        has_enroll = any('enrollment' in t for t in row_tokens)
        has_name = any('student name' in t for t in row_tokens)
        if has_enroll and has_name:
            header_row = i
            break

    if header_row is None:
        return None, 'Could not detect marksheet header row (Enrollment No / Student Name).'

    data_start = header_row + 1
    header_block = raw_df.iloc[:data_start].copy()
    header_block = header_block.fillna('')
    header_block = header_block.replace('', pd.NA).ffill(axis=1).fillna('')

    col_texts = {}
    for col in header_block.columns:
        bits = []
        for row_idx in header_block.index:
            piece = _normalize_text(header_block.iat[row_idx, col])
            if piece:
                bits.append(piece)
        col_texts[col] = ' | '.join(dict.fromkeys(bits))

    def find_col(*needles):
        for col, text in col_texts.items():
            if all(n in text for n in needles):
                return col
        return None

    enrollment_col = find_col('enrollment')
    student_name_col = find_col('student', 'name')
    result_col = find_col('result')
    spi_col = find_col('spi')
    cpi_col = find_col('cpi')
    current_credit_col = find_col('current semester', 'cr.er')
    current_gp_col = find_col('current semester', 'gp')
    cumulative_credit_col = find_col('cummulative', 'cr.er') or find_col('cumulative', 'cr.er')
    cumulative_gp_col = find_col('cummulative', 'gp') or find_col('cumulative', 'gp')

    if enrollment_col is None or student_name_col is None:
        return None, 'Could not detect Enrollment No and Student Name columns.'

    if semester_no is None:
        semester_col = find_col('semester') or find_col('sem')
        if semester_col is not None:
            for row_idx in range(data_start, len(raw_df.index)):
                candidate = _parse_number(raw_df.iat[row_idx, semester_col], as_int=True)
                if candidate and 1 <= candidate <= 8:
                    semester_no = candidate
                    break

    if semester_no is None:
        return None, 'Could not detect semester from marksheet header.'

    if not exam_session:
        exam_session = f'SEM {semester_no}'

    summary_start = result_col if result_col is not None else len(raw_df.columns)
    if current_credit_col is not None:
        summary_start = min(summary_start, current_credit_col)
    if current_gp_col is not None:
        summary_start = min(summary_start, current_gp_col)
    if cumulative_credit_col is not None:
        summary_start = min(summary_start, cumulative_credit_col)
    if cumulative_gp_col is not None:
        summary_start = min(summary_start, cumulative_gp_col)
    if spi_col is not None:
        summary_start = min(summary_start, spi_col)
    if cpi_col is not None:
        summary_start = min(summary_start, cpi_col)

    subject_columns = []
    for col in range(student_name_col + 1, summary_start):
        text = col_texts.get(col, '')
        if not re.search(r'(^|\W)gp(\W|$)', text):
            continue
        if 'current semester' in text or 'cummulative' in text or 'cumulative' in text:
            continue

        code_match = re.search(r'\b[A-Z]{2,}\d{2,}[A-Z0-9\-()]*\b', text.upper())
        if not code_match:
            continue

        code = code_match.group(0)
        cleaned = text
        cleaned = cleaned.replace(code.lower(), ' ')
        cleaned = re.sub(r'(^|\W)gp(\W|$)', ' ', cleaned)
        cleaned = re.sub(r'(^|\W)rem\.?($|\W)', ' ', cleaned)
        cleaned = re.sub(r'\|', ' ', cleaned)
        cleaned = re.sub(r'\s+', ' ', cleaned).strip(' -')
        subject_columns.append({
            'col': col,
            'course_code': code,
            'parsed_name': cleaned.title() if cleaned else code,
        })

    if not subject_columns:
        return None, 'Could not detect any subject GP columns from consolidated marksheet.'

    subject_lookup = {
        s.code.upper(): s
        for s in Subject.objects.select_related('semester').all()
    }

    rows = []
    blanks_after_data = 0
    for row_idx in range(data_start, len(raw_df.index)):
        enrollment = _clean_cell(raw_df.iat[row_idx, enrollment_col])
        if not enrollment or not _looks_like_enrollment(enrollment):
            if rows:
                blanks_after_data += 1
                if blanks_after_data > 20:
                    break
            continue

        blanks_after_data = 0
        result_status = _clean_cell(raw_df.iat[row_idx, result_col]) if result_col is not None else 'PASS'
        if not result_status:
            result_status = 'PASS'

        for subject_col in subject_columns:
            grade = _clean_cell(raw_df.iat[row_idx, subject_col['col']]).upper()
            if not grade:
                continue

            course_code = subject_col['course_code'].upper()
            subject_obj = subject_lookup.get(course_code)
            course_name = subject_obj.name if subject_obj else subject_col['parsed_name']
            course_credit = float(subject_obj.credit) if subject_obj else 0.0

            rows.append({
                'enrollment_no': enrollment,
                'semester': semester_no,
                'exam_session': exam_session,
                'issued_date': None,
                'spi': _parse_number(raw_df.iat[row_idx, spi_col]) if spi_col is not None else None,
                'cpi': _parse_number(raw_df.iat[row_idx, cpi_col]) if cpi_col is not None else None,
                'earned_credits': _parse_number(raw_df.iat[row_idx, current_credit_col], as_int=True) if current_credit_col is not None else None,
                'earned_grade_points': _parse_number(raw_df.iat[row_idx, current_gp_col], as_int=True) if current_gp_col is not None else None,
                'total_credits': _parse_number(raw_df.iat[row_idx, cumulative_credit_col], as_int=True) if cumulative_credit_col is not None else None,
                'total_grade_points': _parse_number(raw_df.iat[row_idx, cumulative_gp_col], as_int=True) if cumulative_gp_col is not None else None,
                'result_status': result_status,
                'course_code': course_code,
                'course_name': course_name,
                'course_credit': course_credit,
                'grade': grade,
            })

    if not rows:
        return None, 'No student grade rows found in consolidated marksheet.'

    return rows, None


def _has_webpush_config():
    return bool(settings.WEBPUSH_PUBLIC_KEY and settings.WEBPUSH_PRIVATE_KEY and webpush)


def _webpush_config_reason():
    if not webpush:
        return 'pywebpush is not installed on server.'
    if not settings.WEBPUSH_PUBLIC_KEY or not settings.WEBPUSH_PRIVATE_KEY:
        return 'WEBPUSH_PUBLIC_KEY / WEBPUSH_PRIVATE_KEY is missing.'
    return ''


def _send_push_notification(title, body, target_url):
    if not _has_webpush_config():
        return 0

    payload = json.dumps({
        'title': title,
        'body': body,
        'url': target_url,
    })

    subscriptions = PushSubscription.objects.filter(is_active=True)
    sent_count = 0
    for subscription in subscriptions:
        try:
            webpush(
                subscription_info={
                    'endpoint': subscription.endpoint,
                    'keys': {
                        'p256dh': subscription.p256dh,
                        'auth': subscription.auth,
                    },
                },
                data=payload,
                vapid_private_key=settings.WEBPUSH_PRIVATE_KEY,
                vapid_claims={
                    'sub': settings.WEBPUSH_SUBJECT,
                },
            )
            sent_count += 1
        except WebPushException as exc:
            status_code = getattr(getattr(exc, 'response', None), 'status_code', None)
            if status_code in (404, 410):
                subscription.is_active = False
                subscription.save(update_fields=['is_active', 'updated_at'])

    return sent_count


def webpush_public_key(request):
    return JsonResponse({
        'publicKey': settings.WEBPUSH_PUBLIC_KEY,
        'configured': _has_webpush_config(),
        'reason': _webpush_config_reason(),
    })


@require_POST
def webpush_subscribe(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Invalid payload.'}, status=400)

    subscription = data.get('subscription') or {}
    endpoint = (subscription.get('endpoint') or '').strip()
    keys = subscription.get('keys') or {}
    p256dh = (keys.get('p256dh') or '').strip()
    auth_key = (keys.get('auth') or '').strip()

    if not endpoint or not p256dh or not auth_key:
        return JsonResponse({'ok': False, 'error': 'Invalid subscription object.'}, status=400)

    user_agent = request.META.get('HTTP_USER_AGENT', '')[:300]
    PushSubscription.objects.update_or_create(
        endpoint=endpoint,
        defaults={
            'p256dh': p256dh,
            'auth': auth_key,
            'user_agent': user_agent,
            'is_active': True,
        },
    )
    return JsonResponse({'ok': True})


@require_POST
def webpush_unsubscribe(request):
    try:
        data = json.loads(request.body.decode('utf-8'))
    except Exception:
        return JsonResponse({'ok': False, 'error': 'Invalid payload.'}, status=400)

    endpoint = (data.get('endpoint') or '').strip()
    if not endpoint:
        return JsonResponse({'ok': False, 'error': 'Endpoint is required.'}, status=400)

    PushSubscription.objects.filter(endpoint=endpoint).update(is_active=False)
    return JsonResponse({'ok': True})


def service_worker(request):
    script = """
self.addEventListener('push', function(event) {
  let payload = { title: 'New Update', body: 'New announcement available.', url: '/announcements/' };
  try {
    payload = event.data ? event.data.json() : payload;
  } catch (err) {}

  event.waitUntil(
    self.registration.showNotification(payload.title || 'New Update', {
      body: payload.body || 'New announcement available.',
      icon: '/static/allocation/images/college_logo.png',
      badge: '/static/allocation/images/college_logo.png',
      data: { url: payload.url || '/announcements/' }
    })
  );
});

self.addEventListener('notificationclick', function(event) {
  event.notification.close();
  const target = (event.notification.data && event.notification.data.url) || '/announcements/';
  event.waitUntil(clients.openWindow(target));
});
""".strip()
    return HttpResponse(script, content_type='application/javascript')


def _create_result_announcement(sheet_map, created_count, updated_count, user):
    if not sheet_map:
        return

    semester_numbers = sorted({data['semester'].number for data in sheet_map.values()})
    exam_sessions = sorted({data['exam_session'] for data in sheet_map.values()})
    total_sheets = len(sheet_map)

    sem_text = ', '.join(f"Sem {sem}" for sem in semester_numbers)
    session_text = ', '.join(exam_sessions)
    title = f"Result Declared: {session_text}"
    body = (
        f"Results are now available for {sem_text}. "
        f"Total records processed: {total_sheets} "
        f"(new: {created_count}, updated: {updated_count}). "
        f"Please check the Result Lookup page."
    )

    source_raw = '|'.join(f"{session}:{sem}" for session in exam_sessions for sem in semester_numbers)
    source_hash = hashlib.md5(source_raw.encode('utf-8')).hexdigest()
    source_key = f"result:{source_hash}"

    notice, _ = Notice.objects.update_or_create(
        source_key=source_key,
        defaults={
            'notice_type': 'RESULT',
            'title': title,
            'body': body,
            'is_published': True,
            'published_at': timezone.now(),
            'created_by': user,
        },
    )
    if notice.is_published:
        _send_push_notification(
            title=notice.title,
            body=notice.body[:180],
            target_url='/announcements/?type=result',
        )


@login_required(login_url='login')
def download_student_upload_sample(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="students_upload_sample.csv"'
    writer = csv.writer(response)
    writer.writerow(['enrollment_no', 'name', 'division', 'admission_year', 'mentor_name', 'email', 'phone', 'elective'])
    writer.writerow(['20BCE30149', 'PATEL HET PARESHKUMAR', 'A', '2020', 'Prof. N Shah', 'het.patel@example.com', '9876543210', 'CT704A-N'])
    writer.writerow(['20BCE30150', 'PATEL HETVI JAGDISH', 'A', '2020', 'Prof. R Patel', 'hetvi.patel@example.com', '9876501234', 'CT704C-N'])
    return response


@login_required(login_url='login')
def download_attendance_upload_sample(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_students_upload_sample.csv"'
    writer = csv.writer(response)
    writer.writerow(['Enrollment No', 'Name', 'Division', 'Elective', 'Room No', 'Exam No'])
    writer.writerow(['20BCE30149', 'PATEL HET PARESHKUMAR', 'A', 'CT703A', '101', '1201'])
    writer.writerow(['20BCE30150', 'PATEL HETVI JAGDISH', 'A', 'CT703C', '101', '1202'])
    return response


@login_required(login_url='login')
def download_marks_upload_sample(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    layout = (request.GET.get('layout') or 'flat').strip().lower()

    if layout == 'consolidated':
        try:
            from openpyxl import Workbook
        except Exception:
            return HttpResponse('openpyxl is required for consolidated sample export.', status=500)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'Consolidated Sample'

        worksheet.append([
            'CONSOLIDATED STATEMENT OF MARKS FOR BE Computer Engineering SEMESTER : VII OCTOBER 2023 EXAMINATION'
        ])
        headers = [
            'EXAM NO.',
            'ENROLLMENT NO.',
            'STUDENT NAME',
            'CT701-N (5) COMPILER DESIGN GP',
            'CT702-N (4) CYBER SECURITY GP',
            'CT703A-N (5) HIGH PERFORMANCE COMPUTING GP',
            'CT703C-N (5) NATURAL LANGUAGE PROCESSING GP',
            'CT703D-N (5) BLOCKCHAIN TECHNOLOGY GP',
            'CT704A-N (5) DISTRIBUTED SYSTEMS GP',
            'CT704C-N (5) IMAGE PROCESSING GP',
            'CT705-N (5) PROJECT-II GP',
            'CURRENT SEMESTER CR.ER.',
            'CURRENT SEMESTER GP',
            'SPI',
            'CUMMULATIVE CR.ER.',
            'CUMMULATIVE GP',
            'CPI',
            'RESULT',
        ]
        worksheet.append(headers)
        worksheet.append([
            '227140',
            '20BCE30149',
            'PATEL HET PARESHKUMAR',
            'A-', 'A-', 'B+', 'A', 'A-', 'A-', 'A', 'A',
            24, 197, 8.21, 170, 1420, 8.35, 'PASS',
        ])
        worksheet.append([
            '227141',
            '20BCE30150',
            'PATEL HETVI JAGDISH',
            'A', 'A', 'A-', 'A', 'A', 'A-', 'A', 'A',
            24, 206, 8.58, 170, 1454, 8.55, 'PASS',
        ])

        stream = io.BytesIO()
        workbook.save(stream)
        stream.seek(0)
        response = HttpResponse(
            stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="marks_upload_consolidated_sample.xlsx"'
        return response

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="marks_upload_flat_sample.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'enrollment_no', 'semester', 'exam_session', 'issued_date', 'spi', 'cpi',
        'earned_credits', 'earned_grade_points', 'total_credits', 'total_grade_points',
        'result_status', 'course_code', 'course_name', 'course_credit', 'grade'
    ])
    writer.writerow([
        '20BCE30149', 7, 'OCTOBER 2023', '2023-12-20', 8.21, 8.35,
        24, 197, 170, 1420, 'PASS', 'CT701-N', 'COMPILER DESIGN', 5, 'A-'
    ])
    writer.writerow([
        '20BCE30149', 7, 'OCTOBER 2023', '2023-12-20', 8.21, 8.35,
        24, 197, 170, 1420, 'PASS', 'CT702-N', 'CYBER SECURITY', 4, 'A-'
    ])
    return response


@login_required(login_url='login')
def download_elective_upload_sample(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="elective_upload_sample.csv"'
    writer = csv.writer(response)
    writer.writerow(['enrollment_no', 'elective_group', 'subject_code'])
    writer.writerow(['20BCE30149', 'ELECTIVE-I', 'CT704A-N'])
    writer.writerow(['20BCE30150', 'ELECTIVE-I', 'CT704C-N'])
    return response


@login_required(login_url='login')
def download_seating_pdf_sample(request):
    if not request.user.is_staff and not hasattr(request.user, 'faculty'):
        return redirect('dashboard')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="seating_summary_sample.pdf"'
    pdf = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    pdf.setFont('Helvetica-Bold', 12)
    pdf.drawCentredString(width / 2, height - 50, 'SEATING SUMMARY SAMPLE')
    pdf.setFont('Helvetica', 10)
    pdf.drawString(40, height - 80, 'Use this numeric block style so parser can read room allocation from PDF:')
    pdf.drawString(40, height - 98, 'Format: <Block No> <Room No> <Seat From> <Seat To> <Count>')

    y = height - 130
    sample_lines = [
        '1 101 227140 227170 31',
        '2 102 227171 227200 30',
        '3 103 227201 227230 30',
    ]
    for line in sample_lines:
        pdf.drawString(60, y, line)
        y -= 20

    pdf.setFont('Helvetica-Oblique', 9)
    pdf.drawString(40, y - 10, 'Numbers are intentionally plain for reliable parsing.')
    pdf.save()
    return response

# ============= PUBLIC HOME & RESULT HELPERS =============

def public_home(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return render(request, 'home.html')

def _build_result_context(student, semester_no=None):
    sheet_qs = ResultSheet.objects.filter(student=student)
    if semester_no:
        sheet_qs = sheet_qs.filter(semester__number=semester_no)
    sheet = sheet_qs.order_by('-created_at').first()

    if not sheet:
        return None, 'Result not available yet. Please contact Exam Cell.'

    entries = sheet.entries.all().order_by('course_code')
    if not entries.exists():
        return None, 'Result data is incomplete. Please contact Exam Cell.'

    rows = []
    for entry in entries:
        rows.append({
            'code': entry.course_code,
            'name': entry.course_name,
            'credit': entry.course_credit,
            'grade': entry.grade,
        })

    return {
        'student': student,
        'sheet': sheet,
        'rows': rows,
    }, None


def _derive_batch_years(enrollment_no):
    prefix = ''.join(ch for ch in (enrollment_no or '') if ch.isdigit())[:2]
    if len(prefix) != 2:
        current_year = timezone.now().year
        return current_year - 4, current_year

    yy = int(prefix)
    current_two_digits = timezone.now().year % 100
    if yy <= current_two_digits + 1:
        start_year = 2000 + yy
    else:
        start_year = 1900 + yy

    return start_year, start_year + 4


def _calculate_attendance_percentage_for_student(student_id):
    agg = Attendance.objects.filter(student_id=student_id).aggregate(
        total=Count('id'),
        present=Sum(Case(When(status='P', then=1), default=0, output_field=IntegerField())),
    )
    total = agg.get('total') or 0
    present = agg.get('present') or 0
    if total <= 0:
        return None
    return round((present * 100.0) / total, 1)


def _compute_student_risk(student):
    attendance_pct = _calculate_attendance_percentage_for_student(student.id)
    marks_qs = StudentMark.objects.filter(student=student).order_by('-updated_at')

    marks_rows = list(marks_qs[:30])
    fail_count = 0
    valid_percentages = []
    for record in marks_rows:
        if record.is_absent:
            fail_count += 1
            continue
        if record.pass_marks is not None and record.marks_obtained is not None and record.marks_obtained < record.pass_marks:
            fail_count += 1
        if record.marks_obtained is not None and record.max_marks and record.max_marks > 0:
            valid_percentages.append(float(record.marks_obtained) * 100.0 / float(record.max_marks))

    avg_marks_pct = round(sum(valid_percentages) / len(valid_percentages), 1) if valid_percentages else None

    score = 0
    suggestions = []

    if attendance_pct is not None:
        if attendance_pct < 75:
            score += 25
            suggestions.append('Attendance improvement plan')
        if attendance_pct < 60:
            score += 15

    if fail_count > 0:
        score += min(40, fail_count * 12)
        suggestions.append('Remedial support recommended')

    if avg_marks_pct is not None:
        if avg_marks_pct < 50:
            score += 20
        elif avg_marks_pct < 65:
            score += 10

    latest = valid_percentages[:5]
    previous = valid_percentages[5:10]
    if latest and previous:
        latest_avg = sum(latest) / len(latest)
        previous_avg = sum(previous) / len(previous)
        if latest_avg + 8 < previous_avg:
            score += 10
            suggestions.append('Marks trend is declining')

    if score >= 60:
        level = 'HIGH'
    elif score >= 35:
        level = 'MEDIUM'
    else:
        level = 'LOW'

    if level == 'HIGH' and 'Mentor counselling needed' not in suggestions:
        suggestions.append('Mentor counselling needed')

    return {
        'score': min(score, 100),
        'level': level,
        'attendance_pct': attendance_pct,
        'avg_marks_pct': avg_marks_pct,
        'fail_count': fail_count,
        'suggestions': suggestions,
    }


def _build_student_timeline(student):
    items = []

    for mark in StudentMark.objects.filter(student=student).select_related('subject', 'semester').order_by('-updated_at')[:40]:
        if mark.is_absent:
            result_label = 'Absent'
        elif mark.marks_obtained is None:
            result_label = 'Pending'
        elif mark.pass_marks is not None and mark.marks_obtained < mark.pass_marks:
            result_label = 'Fail'
        else:
            result_label = 'Pass'

        items.append({
            'date': mark.updated_at,
            'type': 'MARK',
            'title': f"{mark.subject.code} {mark.get_exam_type_display()} Attempt {mark.attempt_no}",
            'meta': f"{mark.exam_session} | {result_label}",
        })

    for log in MentorActionLog.objects.filter(student=student).select_related('created_by').order_by('-created_at')[:20]:
        author = log.created_by.get_full_name() if log.created_by else 'System'
        items.append({
            'date': log.created_at,
            'type': 'MENTOR',
            'title': log.get_action_type_display(),
            'meta': f"{author}: {log.note[:120]}",
        })

    for sheet in ResultSheet.objects.filter(student=student).select_related('semester').order_by('-created_at')[:10]:
        items.append({
            'date': sheet.created_at,
            'type': 'RESULT',
            'title': f"Result uploaded - Sem {sheet.semester.number}",
            'meta': f"{sheet.exam_session} | {sheet.result_status}",
        })

    items.sort(key=lambda row: row['date'], reverse=True)
    return items[:50]


def _get_mentor_scope(user):
    if getattr(user, 'is_staff', False):
        return None
    faculty = getattr(user, 'faculty', None)
    if not faculty:
        return set()
    return set(
        MentorAssignment.objects.filter(faculty=faculty).values_list('semester_id', 'division')
    )


def _is_class_mentor(user, student):
    if getattr(user, 'is_staff', False):
        return True
    faculty = getattr(user, 'faculty', None)
    if not faculty or not student:
        return False
    return MentorAssignment.objects.filter(
        faculty=faculty,
        semester=student.semester,
        division=student.division,
    ).exists()


def _is_mentor_for_scope(user, semester_obj, divisions):
    if getattr(user, 'is_staff', False):
        return True

    faculty = getattr(user, 'faculty', None)
    if not faculty or not semester_obj:
        return False

    divisions_to_check = list(divisions or [])
    if not divisions_to_check:
        return False

    assigned = set(
        MentorAssignment.objects.filter(
            faculty=faculty,
            semester=semester_obj,
            division__in=divisions_to_check,
        ).values_list('division', flat=True)
    )
    return set(divisions_to_check).issubset(assigned)


def parse_rooms(room_text):
    rooms = []
    seen = set()

    for part in room_text.split(','):
        part = part.strip()

        if '-' in part:
            start, end = part.split('-')
            for r in range(int(start), int(end) + 1):
                room = str(r)
                if room not in seen:
                    rooms.append(room)
                    seen.add(room)
        else:
            if part and part not in seen:
                rooms.append(part)
                seen.add(part)

    return rooms


def _parse_seating_pdf(uploaded_pdf):
    """Parse seating summary PDF to extract room seat ranges and counts."""
    try:
        from PyPDF2 import PdfReader
    except Exception:
        return [], "PyPDF2 is not installed."

    try:
        reader = PdfReader(uploaded_pdf)
        text = "\n".join([page.extract_text() or "" for page in reader.pages])
    except Exception as exc:
        return [], f"Failed to read PDF: {exc}"

    room_specs = []
    seen = set()
    tokens = re.findall(r"\d+", text)

    for i in range(len(tokens) - 3):
        block_str = tokens[i]
        room_str = tokens[i + 1]
        seat_from_str = tokens[i + 2]
        seat_to_str = tokens[i + 3]

        if len(seat_from_str) < 5 or len(seat_to_str) < 5:
            continue

        try:
            block_no = int(block_str)
            room_no = int(room_str)
            seat_from = int(seat_from_str)
            seat_to = int(seat_to_str)
        except ValueError:
            continue

        if block_no < 1 or block_no > 200:
            continue
        if room_no < 1 or room_no > 999:
            continue
        if seat_to < seat_from:
            continue

        count = None
        if i + 4 < len(tokens):
            try:
                count_candidate = int(tokens[i + 4])
            except ValueError:
                count_candidate = None
            if count_candidate == seat_to - seat_from + 1:
                count = count_candidate

        if count is None:
            count = seat_to - seat_from + 1

        if room_no in seen:
            continue

        room_specs.append({
            'room_no': str(room_no),
            'seat_from': seat_from,
            'seat_to': seat_to,
            'count': count,
        })
        seen.add(room_no)

    if not room_specs:
        return [], "No room data found in PDF."
    return room_specs, None


def _parse_attendance_students_file(uploaded_file):
    """Parse uploaded CSV/Excel for attendance report rows."""
    filename = (uploaded_file.name or '').lower()
    is_csv = filename.endswith('.csv')
    is_legacy_xls = filename.endswith('.xls')
    is_excel = filename.endswith(('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'))

    if not is_csv and not is_excel:
        return [], "Unsupported file type. Please upload a .csv or .xlsx file."

    try:
        if is_csv:
            df = pd.read_csv(uploaded_file)
        elif is_legacy_xls:
            df = pd.read_excel(uploaded_file)
        else:
            try:
                import openpyxl  # noqa: F401
            except Exception:
                return [], (
                    "Could not read uploaded student file: openpyxl is not installed in the "
                    f"Django server environment. Run '{sys.executable} -m pip install openpyxl' "
                    "and restart the server."
                )
            df = pd.read_excel(uploaded_file, engine='openpyxl')
    except Exception as exc:
        return [], f"Could not read uploaded student file: {exc}"

    if df is None or df.empty:
        return [], "Uploaded student file is empty."

    def normalize_col(value):
        return re.sub(r'[^a-z0-9]+', '', str(value).strip().lower())

    column_map = {normalize_col(col): col for col in df.columns}

    def find_col(*aliases):
        for alias in aliases:
            col = column_map.get(alias)
            if col is not None:
                return col
        return None

    enrollment_col = find_col('enrollmentno', 'enrollment', 'enrolmentno', 'enrolment')
    name_col = find_col('name', 'studentname', 'student')
    division_col = find_col('division', 'div')
    elective_col = find_col('elective', 'electives', 'electivesubject')
    elective1_col = find_col('elective1', 'elective_1', 'electiveone', 'electivefirst')
    elective2_col = find_col('elective2', 'elective_2', 'electivesecond')
    room_no_col = find_col('roomno', 'roomnumber', 'room')
    exam_no_col = find_col('examno', 'examnumber', 'seatno')

    if enrollment_col is None or name_col is None:
        return [], (
            "Invalid file format. Required columns: Enrollment No and Name "
            "(case-insensitive). Optional: Division, Elective, Room No, Exam No."
        )

    rows = []
    skipped = 0
    for _, row in df.iterrows():
        enrollment_no = _clean_cell(row.get(enrollment_col))
        name = _clean_cell(row.get(name_col))

        if not enrollment_no and not name:
            continue
        if not enrollment_no:
            skipped += 1
            continue

        elective_1 = _clean_cell(row.get(elective1_col)) if elective1_col is not None else ''
        elective_2 = _clean_cell(row.get(elective2_col)) if elective2_col is not None else ''
        if not elective_1 and not elective_2 and elective_col is not None:
            elective_raw = _clean_cell(row.get(elective_col))
            if elective_raw:
                parts = [
                    part.strip()
                    for part in re.split(r'[;,|]+', elective_raw)
                    if part and part.strip()
                ]
                elective_1 = parts[0] if len(parts) > 0 else ''
                elective_2 = parts[1] if len(parts) > 1 else ''

        rows.append({
            'sr': 0,
            'enrollment_no': enrollment_no,
            'name': name,
            'division': _clean_cell(row.get(division_col)) if division_col is not None else '',
            'elective_1': elective_1,
            'elective_2': elective_2,
            'room_no': _clean_cell(row.get(room_no_col)) if room_no_col is not None else '',
            'exam_no': _clean_cell(row.get(exam_no_col)) if exam_no_col is not None else '',
        })

    if not rows:
        if skipped:
            return [], "No valid rows found. Every row must include Enrollment No."
        return [], "No student rows found in uploaded file."

    return rows, None

from datetime import datetime, timedelta

# ============= DASHBOARD VIEWS =============

@login_required(login_url='login')
def dashboard(request):
    """Main dashboard - redirects based on role"""
    if request.user.is_staff:
        return admin_dashboard(request)
    elif hasattr(request.user, 'faculty'):
        return faculty_dashboard(request)
    else:
        # Break redirect loop: show simple access message instead of redirecting
        return render(request, 'auth/login.html', {
            'form': None,
            'error': 'Your account is not assigned a role. Please contact admin.'
        })

def admin_dashboard(request):
    """Admin (Exam Cell) Dashboard"""
    semesters = list(Semester.objects.all().order_by('number'))
    division_codes = [code for code, _ in Student.DIVISION_CHOICES]

    student_counts = Student.objects.values('semester_id', 'division').annotate(total=Count('id'))
    student_totals = Student.objects.values('semester_id').annotate(total=Count('id'))
    subject_counts = Subject.objects.values('semester_id').annotate(total=Count('id'))
    subject_elective_counts = Subject.objects.filter(is_elective=True).values('semester_id').annotate(total=Count('id'))

    student_lookup = {(row['semester_id'], row['division']): row['total'] for row in student_counts}
    student_total_lookup = {row['semester_id']: row['total'] for row in student_totals}
    subject_lookup = {row['semester_id']: row['total'] for row in subject_counts}
    subject_elective_lookup = {row['semester_id']: row['total'] for row in subject_elective_counts}

    semester_cards = []
    for sem in semesters:
        division_rows = []
        for div in division_codes:
            division_rows.append({
                'division': div,
                'count': student_lookup.get((sem.id, div), 0),
            })
        total_subjects = subject_lookup.get(sem.id, 0)
        elective_subjects = subject_elective_lookup.get(sem.id, 0)
        compulsory_subjects = max(total_subjects - elective_subjects, 0)
        semester_cards.append({
            'semester': sem.number,
            'student_total': student_total_lookup.get(sem.id, 0),
            'subject_count': total_subjects,
            'subject_elective_count': elective_subjects,
            'subject_compulsory_count': compulsory_subjects,
            'division_rows': division_rows,
        })

    context = {
        'semester_cards': semester_cards,
    }
    return render(request, 'dashboard/admin_dashboard.html', context)


@login_required(login_url='login')
def sample_files(request):
    """Central hub for sample upload formats."""
    if not request.user.is_staff and not hasattr(request.user, 'faculty'):
        return redirect('dashboard')

    return render(request, 'dashboard/sample_files.html', {
        'is_admin': request.user.is_staff,
    })


@login_required(login_url='login')
def manage_notices(request):
    if not request.user.is_staff:
        return redirect('dashboard')

    message = None
    active_subscriptions = PushSubscription.objects.filter(is_active=True).count()
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip().lower()
        if action == 'send_test_push':
            form = NoticeForm(initial={'is_published': True, 'notice_type': 'NOTICE'})
            if not _has_webpush_config():
                message = 'Web push is not configured. Add WEBPUSH keys in environment first.'
            elif active_subscriptions == 0:
                message = 'No subscribed devices found yet. Open Announcements page on a device and allow notifications.'
            else:
                delivered = _send_push_notification(
                    title='Test Notification',
                    body='Push notifications are working for notices and results.',
                    target_url='/announcements/',
                )
                message = f'Test notification sent to {delivered} device(s).'
        elif action == 'delete_attachment':
            attachment_id = request.POST.get('attachment_id')
            attachment = NoticeAttachment.objects.filter(id=attachment_id).select_related('notice').first()
            if not attachment:
                message = 'Attachment not found.'
            else:
                if attachment.file and default_storage.exists(attachment.file.name):
                    default_storage.delete(attachment.file.name)
                attachment.delete()
                message = 'Attachment deleted successfully.'
            form = NoticeForm(initial={'is_published': True, 'notice_type': 'NOTICE'})
        elif action == 'delete_notice':
            notice_id = request.POST.get('notice_id')
            notice = Notice.objects.filter(id=notice_id).first()
            if not notice:
                message = 'Notice not found.'
            else:
                notice_title = notice.title
                for attachment in notice.attachments.all():
                    if attachment.file and default_storage.exists(attachment.file.name):
                        default_storage.delete(attachment.file.name)
                notice.delete()
                message = f'Notice "{notice_title}" deleted successfully.'
            form = NoticeForm(initial={'is_published': True, 'notice_type': 'NOTICE'})
        else:
            form = NoticeForm(request.POST)
            if form.is_valid():
                notice = form.save(commit=False)
                notice.created_by = request.user
                if notice.is_published and not notice.published_at:
                    notice.published_at = timezone.now()
                if not notice.is_published:
                    notice.published_at = None
                notice.save()
                uploaded_files = request.FILES.getlist('attachments')
                for uploaded_file in uploaded_files:
                    NoticeAttachment.objects.create(
                        notice=notice,
                        file=uploaded_file,
                        file_name=uploaded_file.name,
                    )
                message = 'Notice published successfully.' if notice.is_published else 'Notice saved as draft.'
                if uploaded_files:
                    message += f' {len(uploaded_files)} attachment(s) uploaded.'
                if notice.is_published:
                    target_url = '/announcements/?type=result' if notice.notice_type == 'RESULT' else '/announcements/?type=notice'
                    _send_push_notification(
                        title=notice.title,
                        body=notice.body[:180],
                        target_url=target_url,
                    )
                form = NoticeForm(initial={'is_published': True, 'notice_type': 'NOTICE'})
    else:
        form = NoticeForm(initial={'is_published': True, 'notice_type': 'NOTICE'})

    notices = Notice.objects.select_related('created_by').prefetch_related('attachments').all()[:100]
    return render(request, 'admin/manage_notices.html', {
        'form': form,
        'message': message,
        'notices': notices,
        'active_subscriptions': active_subscriptions,
        'webpush_configured': _has_webpush_config(),
    })


@ensure_csrf_cookie
def public_notices(request):
    selected_type = (request.GET.get('type') or 'all').strip().lower()
    notices = Notice.objects.filter(is_published=True).prefetch_related('attachments')

    if selected_type == 'notice':
        notices = notices.filter(notice_type='NOTICE')
    elif selected_type == 'result':
        notices = notices.filter(notice_type='RESULT')
    else:
        selected_type = 'all'

    return render(request, 'results/public_notices.html', {
        'notices': notices[:200],
        'selected_type': selected_type,
        'new_cutoff': timezone.now() - timedelta(days=7),
    })


@login_required(login_url='login')
def seating_home(request):
    """Seating landing page with options."""
    if not request.user.is_staff and not hasattr(request.user, 'faculty'):
        return redirect('dashboard')
    return render(request, 'seating/home.html')

def faculty_dashboard(request):
    """Faculty Dashboard"""
    try:
        faculty = request.user.faculty
    except:
        return redirect('login')
    
    subjects = faculty.subjects.all()
    mentored_classes = MentorAssignment.objects.filter(faculty=faculty).select_related('semester').order_by('semester__number', 'division')
    
    # Get last accessed subject/semester
    last_subject = faculty.last_subject
    last_semester = faculty.last_semester
    
    # Pending work indicators
    pending_attendance = {}
    
    for subject in subjects:
        # Check attendance
        today = timezone.now().date()
        today_attendance = Attendance.objects.filter(
            subject=subject,
            date=today,
            marked_by=faculty
        ).count()
        
        if today_attendance == 0:
            pending_attendance[subject.id] = 'Not marked today'
        
        # Marks entry removed from system
    
    context = {
        'faculty': faculty,
        'subjects': subjects,
        'last_subject': last_subject,
        'last_semester': last_semester,
        'pending_attendance': pending_attendance,
        'mentored_classes': mentored_classes,
        # 'pending_marks' removed
    }
    return render(request, 'dashboard/faculty_dashboard.html', context)

# ============= STUDENT MANAGEMENT =============

@login_required(login_url='login')
def student_list(request):
    """List and search students"""
    # Allow both admin and faculty to view students
    
    form = StudentSearchForm(request.GET)
    students = Student.objects.all()
    
    selected_divisions = []
    selected_risk_level = (request.GET.get('risk_level') or '').strip().upper()
    if request.GET:
        query = request.GET.get('query', '')
        branch = request.GET.get('branch')
        semester = request.GET.get('semester')
        selected_divisions = [d for d in request.GET.getlist('division') if d]
        if not selected_divisions:
            single_division = request.GET.get('division')
            if single_division:
                selected_divisions = [single_division]
        mentor = request.GET.get('mentor')
        admission_year = request.GET.get('admission_year')
        sort = request.GET.get('sort', '')
        order = request.GET.get('order', 'asc')
        
        if query:
            students = students.filter(
                Q(name__icontains=query) |
                Q(enrollment_no__icontains=query)
            )
        
        if branch:
            students = students.filter(branch_id=branch)
        
        if semester:
            students = students.filter(semester_id=semester)

        if selected_divisions:
            students = students.filter(division__in=selected_divisions)

        if mentor:
            students = students.filter(mentor_name__icontains=mentor)

        if admission_year:
            students = students.filter(admission_year=admission_year)

        # Sorting
        sort_map = {
            'name': 'name',
            'enrollment_no': 'enrollment_no',
            'branch': 'branch__name',
            'semester': 'semester__number',
            'division': 'division',
            'admission_year': 'admission_year',
            'mentor': 'mentor_name',
            'uploaded': 'id',
        }
        sort_field = sort_map.get(sort, 'enrollment_no')
        if order == 'desc':
            sort_field = '-' + sort_field
        students = students.order_by(sort_field)

    mentor_scope = _get_mentor_scope(request.user)
    student_rows = []
    for student in students:
        risk = _compute_student_risk(student)
        if selected_risk_level and risk['level'] != selected_risk_level:
            continue
        is_mentor_class = True if request.user.is_staff else ((student.semester_id, student.division) in mentor_scope)
        student_rows.append({
            'student': student,
            'risk': risk,
            'is_mentor_class': is_mentor_class,
        })

    if request.GET.get('download') == 'csv':
        export_rows = student_rows
        if not request.user.is_staff:
            export_rows = [row for row in student_rows if row['is_mentor_class']]

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="students.csv"'
        writer = csv.writer(response)
        writer.writerow([
            'Enrollment No', 'Name', 'Branch', 'Semester',
            'Division', 'Admission Year', 'Mentor', 'Email', 'Phone'
        ])
        for row in export_rows:
            student = row['student']
            writer.writerow([
                student.enrollment_no,
                student.name,
                student.branch.name if student.branch else '',
                student.semester.number if student.semester else '',
                student.division,
                student.admission_year or '',
                student.mentor_name or '',
                student.email or '',
                student.phone or '',
            ])
        return response
    
    context = {
        'students': student_rows,
        'form': form,
        'selected_divisions': selected_divisions,
        'selected_risk_level': selected_risk_level,
        'faculty_can_manage_all': request.user.is_staff,
    }
    return render(request, 'students/student_list.html', context)

@login_required(login_url='login')
def student_profile(request, student_id):
    """Student Profile Page - shows complete academic information"""
    student = get_object_or_404(Student, id=student_id)
    can_manage_class = _is_class_mentor(request.user, student)

    if request.method == 'POST' and (request.user.is_staff or hasattr(request.user, 'faculty')):
        action = (request.POST.get('action') or '').strip()
        if action == 'add_mentor_action':
            if not can_manage_class:
                return HttpResponse('Mentor-only action for this class.', status=403)
            action_type = (request.POST.get('action_type') or 'OTHER').strip().upper()
            note = (request.POST.get('note') or '').strip()
            if note:
                MentorActionLog.objects.create(
                    student=student,
                    action_type=action_type if action_type in dict(MentorActionLog.ACTION_CHOICES) else 'OTHER',
                    note=note,
                    created_by=request.user,
                )
            return redirect('student_profile', student_id=student.id)
    
    # Basic info already in student object
    
    # Get enrolled subjects
    enrolled_subjects = StudentSubject.objects.filter(student=student).select_related('subject')
    
    # Get attendance records
    attendance_summary = {}
    for subject_enrollment in enrolled_subjects:
        subject = subject_enrollment.subject
        total_classes = Attendance.objects.filter(subject=subject).count()
        present_count = Attendance.objects.filter(
            student=student,
            subject=subject,
            status='P'
        ).count()
        
        percentage = (present_count / total_classes * 100) if total_classes > 0 else 0
        attendance_summary[subject.code] = {
            'present': present_count,
            'total': total_classes,
            'percentage': round(percentage, 1)
        }
    
    all_marks_qs = StudentMark.objects.filter(student=student).select_related('subject', 'semester')

    selected_exam_type = (request.GET.get('exam_type') or '').strip().upper()
    selected_exam_session = (request.GET.get('exam_session') or '').strip().upper()
    selected_attempt_no = (request.GET.get('attempt_no') or '').strip()

    marks_qs = all_marks_qs
    if selected_exam_type in ('MID', 'FINAL'):
        marks_qs = marks_qs.filter(exam_type=selected_exam_type)
    if selected_exam_session:
        marks_qs = marks_qs.filter(exam_session__iexact=selected_exam_session)
    if selected_attempt_no:
        try:
            marks_qs = marks_qs.filter(attempt_no=int(selected_attempt_no))
        except ValueError:
            pass

    risk = _compute_student_risk(student)
    timeline = _build_student_timeline(student)
    mentor_actions = MentorActionLog.objects.filter(student=student).select_related('created_by')[:20]
    marks_audits = MarksAuditTrail.objects.filter(student=student).select_related('changed_by', 'subject')[:25]

    context = {
        'student': student,
        'enrolled_subjects': enrolled_subjects,
        'attendance_summary': attendance_summary,
        'marks_records': marks_qs.order_by('-updated_at', 'subject__code'),
        'selected_exam_type': selected_exam_type,
        'selected_exam_session': selected_exam_session,
        'selected_attempt_no': selected_attempt_no,
        'exam_session_options': list(
            all_marks_qs.exclude(exam_session='').values_list('exam_session', flat=True).distinct().order_by('exam_session')
        ),
        'attempt_options': list(
            all_marks_qs.values_list('attempt_no', flat=True).distinct().order_by('attempt_no')
        ),
        'risk': risk,
        'timeline': timeline,
        'mentor_actions': mentor_actions,
        'marks_audits': marks_audits,
        'mentor_action_choices': MentorActionLog.ACTION_CHOICES,
        'can_log_actions': can_manage_class,
        'can_manage_class': can_manage_class,
        'can_download_marks': can_manage_class,
    }
    return render(request, 'students/student_profile.html', context)


@login_required(login_url='login')
def student_marks_download_csv(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if not _is_class_mentor(request.user, student):
        return HttpResponse('You can download marks only for classes you mentor.', status=403)

    marks_records = StudentMark.objects.filter(student=student).select_related('subject', 'semester')

    selected_exam_type = (request.GET.get('exam_type') or '').strip().upper()
    selected_exam_session = (request.GET.get('exam_session') or '').strip().upper()
    selected_attempt_no = (request.GET.get('attempt_no') or '').strip()

    if selected_exam_type in ('MID', 'FINAL'):
        marks_records = marks_records.filter(exam_type=selected_exam_type)
    if selected_exam_session:
        marks_records = marks_records.filter(exam_session__iexact=selected_exam_session)
    if selected_attempt_no:
        try:
            marks_records = marks_records.filter(attempt_no=int(selected_attempt_no))
        except ValueError:
            pass

    marks_records = marks_records.order_by('-updated_at', 'subject__code')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{student.enrollment_no}_marks_history.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Enrollment No', 'Student Name', 'Exam Session', 'Exam Type', 'Attempt',
        'Semester', 'Subject Code', 'Subject Name', 'Marks Obtained', 'Max Marks',
        'Pass Marks', 'Absent', 'Status', 'Updated At'
    ])

    for record in marks_records:
        if record.is_absent:
            status = 'ABSENT'
            marks_obtained = ''
        elif record.pass_marks is not None and record.marks_obtained is not None and record.marks_obtained < record.pass_marks:
            status = 'FAIL'
            marks_obtained = record.marks_obtained
        elif record.marks_obtained is not None:
            status = 'PASS'
            marks_obtained = record.marks_obtained
        else:
            status = ''
            marks_obtained = ''

        writer.writerow([
            student.enrollment_no,
            student.name,
            record.exam_session,
            record.get_exam_type_display(),
            record.attempt_no,
            record.semester.number,
            record.subject.code,
            record.subject.name,
            marks_obtained,
            record.max_marks,
            record.pass_marks if record.pass_marks is not None else '',
            'YES' if record.is_absent else 'NO',
            status,
            timezone.localtime(record.updated_at).strftime('%d-%m-%Y %I:%M %p'),
        ])

    return response


@login_required(login_url='login')
def student_marks_download_pdf(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    if not _is_class_mentor(request.user, student):
        return HttpResponse('You can download marks only for classes you mentor.', status=403)

    marks_records = StudentMark.objects.filter(student=student).select_related('subject', 'semester')

    selected_exam_type = (request.GET.get('exam_type') or '').strip().upper()
    selected_exam_session = (request.GET.get('exam_session') or '').strip().upper()
    selected_attempt_no = (request.GET.get('attempt_no') or '').strip()

    if selected_exam_type in ('MID', 'FINAL'):
        marks_records = marks_records.filter(exam_type=selected_exam_type)
    if selected_exam_session:
        marks_records = marks_records.filter(exam_session__iexact=selected_exam_session)
    if selected_attempt_no:
        try:
            marks_records = marks_records.filter(attempt_no=int(selected_attempt_no))
        except ValueError:
            pass

    marks_records = marks_records.order_by('-updated_at', 'subject__code')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{student.enrollment_no}_marks_history.pdf"'

    pdf = canvas.Canvas(response, pagesize=A4)
    page_width, page_height = A4
    y = page_height - 36

    pdf.setFont('Helvetica-Bold', 14)
    pdf.drawString(30, y, 'Student Marks History')
    y -= 18

    pdf.setFont('Helvetica', 10)
    pdf.drawString(30, y, f'Enrollment: {student.enrollment_no}')
    y -= 14
    pdf.drawString(30, y, f'Name: {student.name}')
    y -= 14
    pdf.drawString(30, y, f'Branch/Semester/Division: {student.branch.name} / S{student.semester.number} / {student.division}')
    y -= 20

    table_data = [[
        'Session', 'Type', 'Att', 'Sem', 'Subject', 'Marks', 'Status'
    ]]

    for record in marks_records:
        if record.is_absent:
            marks_text = 'AB'
            status = 'ABSENT'
        elif record.marks_obtained is not None:
            marks_text = f'{record.marks_obtained}/{record.max_marks}'
            if record.pass_marks is not None and record.marks_obtained < record.pass_marks:
                status = 'FAIL'
            else:
                status = 'PASS'
        else:
            marks_text = '-'
            status = '-'

        table_data.append([
            record.exam_session,
            record.get_exam_type_display(),
            str(record.attempt_no),
            str(record.semester.number),
            f'{record.subject.code}',
            marks_text,
            status,
        ])

    if len(table_data) == 1:
        table_data.append(['-', '-', '-', '-', 'No records', '-', '-'])

    table = Table(table_data, colWidths=[1.15 * inch, 0.95 * inch, 0.45 * inch, 0.45 * inch, 1.85 * inch, 0.95 * inch, 0.75 * inch])
    table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 8),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('ALIGN', (2, 0), (3, -1), 'CENTER'),
        ('ALIGN', (5, 0), (6, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))

    table.wrapOn(pdf, page_width - 60, page_height)
    table_height = min(page_height - 140, 16 * len(table_data))
    table.drawOn(pdf, 30, max(40, y - table_height))

    pdf.setFont('Helvetica', 8)
    pdf.drawRightString(page_width - 30, 20, f'Generated: {timezone.localtime().strftime("%d-%m-%Y %I:%M %p")}')
    pdf.save()
    return response


@login_required(login_url='login')
def add_student(request):
    """Add new student (Admin only)"""
    if not request.user.is_staff:
        return redirect('dashboard')
    
    if request.method == 'POST':
        form = StudentForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    else:
        form = StudentForm()
    
    context = {'form': form}
    return render(request, 'students/add_student.html', context)

@login_required(login_url='login')
def delete_student(request, student_id):
    """Delete a student (Admin only) with confirmation."""
    if not request.user.is_staff:
        return redirect('dashboard')

    student = get_object_or_404(Student, id=student_id)

    if request.method == 'POST':
        confirm = request.POST.get('confirm_enrollment', '').strip()
        if confirm != student.enrollment_no:
            return render(request, 'students/confirm_delete.html', {
                'student': student,
                'error': 'Enrollment number does not match. Deletion cancelled.'
            })
        # Cascade deletes will handle related records
        student.delete()
        return redirect('student_list')

    return render(request, 'students/confirm_delete.html', {
        'student': student
    })

@login_required(login_url='login')
@require_POST
def bulk_delete_students(request):
    """Delete all or selected students (Admin only)."""
    if not request.user.is_staff:
        return redirect('dashboard')

    delete_all = request.POST.get('delete_all') == 'on'
    selected_ids = request.POST.getlist('student_ids')

    if delete_all:
        Student.objects.all().delete()
    elif selected_ids:
        Student.objects.filter(id__in=selected_ids).delete()

    next_url = request.POST.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('student_list')

@login_required(login_url='login')
@require_POST
def bulk_promote_students(request):
    """Promote selected students to the next semester (Admin only)."""
    if not request.user.is_staff:
        return redirect('dashboard')

    selected_ids = request.POST.getlist('student_ids')
    if selected_ids:
        semesters = {s.number: s for s in Semester.objects.all()}
        students = list(Student.objects.filter(id__in=selected_ids).select_related('semester'))
        to_update = []
        promoted_class_pairs = set()
        for student in students:
            current_sem = student.semester
            next_number = (student.semester.number or 0) + 1
            next_sem = semesters.get(next_number)
            if next_sem:
                promoted_class_pairs.add((current_sem.id, next_sem.id, student.division))
                student.semester = next_sem
                to_update.append(student)
        if to_update:
            Student.objects.bulk_update(to_update, ['semester'])

        # Keep mentor continuity for each promoted class (Sem+Division -> next Sem+Division).
        for from_sem_id, to_sem_id, division in promoted_class_pairs:
            source_assignments = MentorAssignment.objects.filter(
                semester_id=from_sem_id,
                division=division,
            ).select_related('faculty')

            target_assignment_qs = MentorAssignment.objects.filter(
                semester_id=to_sem_id,
                division=division,
            )
            target_faculty_ids = set(target_assignment_qs.values_list('faculty_id', flat=True))

            for source_assignment in source_assignments:
                if source_assignment.faculty_id in target_faculty_ids:
                    continue
                if len(target_faculty_ids) >= 2:
                    break
                MentorAssignment.objects.create(
                    faculty=source_assignment.faculty,
                    semester_id=to_sem_id,
                    division=division,
                )
                target_faculty_ids.add(source_assignment.faculty_id)

    next_url = request.POST.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect('student_list')


@login_required(login_url='login')
@require_POST
def download_selected_students(request):
    """Download selected students as CSV (available to admin and faculty)."""
    selected_ids = request.POST.getlist('student_ids')
    if not selected_ids:
        next_url = request.POST.get('next', '')
        if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
            return redirect(next_url)
        return redirect('student_list')

    students = (
        Student.objects.filter(id__in=selected_ids)
        .select_related('branch', 'semester')
        .order_by('enrollment_no')
    )

    if not request.user.is_staff:
        mentor_scope = _get_mentor_scope(request.user)
        students = [s for s in students if (s.semester_id, s.division) in mentor_scope]
        if not students:
            return HttpResponse('No selected students are in your mentor classes.', status=403)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="selected_students.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'Enrollment No', 'Name', 'Branch', 'Semester',
        'Division', 'Admission Year', 'Mentor', 'Email', 'Phone'
    ])
    for student in students:
        writer.writerow([
            student.enrollment_no,
            student.name,
            student.branch.name if student.branch else '',
            student.semester.number if student.semester else '',
            student.division,
            student.admission_year or '',
            student.mentor_name or '',
            student.email or '',
            student.phone or '',
        ])
    return response

# ============= ATTENDANCE VIEWS =============

@login_required(login_url='login')
def mark_attendance(request):
    """Mark daily attendance"""
    try:
        faculty = request.user.faculty
    except:
        return redirect('dashboard')
    
    selected_semester_id = request.GET.get('semester') or request.POST.get('semester')
    mentor_error = None

    if request.method == 'POST':
        form = AttendanceForm(request.POST, selected_semester_id=selected_semester_id)
        if form.is_valid():
            subject = form.cleaned_data['subject']
            date = form.cleaned_data['date']

            # Update faculty context
            faculty.last_subject = subject
            faculty.save()

            selected_divisions = [d for d in request.POST.getlist('division') if d]
            if not selected_divisions:
                single_division = request.POST.get('division')
                if single_division:
                    selected_divisions = [single_division]

            if not _is_mentor_for_scope(request.user, subject.semester, selected_divisions):
                mentor_error = 'You can mark attendance only for classes where you are assigned as mentor.'
                form = AttendanceForm(request.POST, selected_semester_id=selected_semester_id)
                context = {
                    'form': form,
                    'students': [],
                    'current_attendance': {},
                    'date': request.POST.get('date'),
                    'semester': request.POST.get('semester'),
                    'subject': request.POST.get('subject'),
                    'selected_subject_obj': None,
                    'selected_divisions': selected_divisions,
                    'mentor_error': mentor_error,
                }
                return render(request, 'attendance/mark_attendance.html', context)

            # Get students enrolled in this subject (optionally filtered by division)
            student_enrollments = StudentSubject.objects.filter(subject=subject)
            if selected_divisions:
                student_enrollments = student_enrollments.filter(student__division__in=selected_divisions)

            # Process attendance from POST data
            for enrollment in student_enrollments:
                student = enrollment.student
                status_key = f'student_{student.id}'
                status = request.POST.get(status_key, 'P')

                # Update or create attendance record
                Attendance.objects.update_or_create(
                    student=student,
                    subject=subject,
                    date=date,
                    defaults={
                        'status': status,
                        'marked_by': faculty
                    }
                )

            return redirect('mark_attendance')
    else:
        form = AttendanceForm(selected_semester_id=selected_semester_id)
    
    def _get_attendance_students(subject_obj, divisions):
        students_qs = Student.objects.filter(
            branch=subject_obj.branch,
            semester=subject_obj.semester,
        )
        if divisions:
            students_qs = students_qs.filter(division__in=divisions)
        return list(students_qs.order_by('enrollment_no'))

    # If subject and date are selected, show student list
    semester = request.GET.get('semester')
    subject = request.GET.get('subject')
    date = request.GET.get('date')
    selected_divisions = [d for d in request.GET.getlist('division') if d]
    if not selected_divisions:
        single_division = request.GET.get('division')
        if single_division:
            selected_divisions = [single_division]
    students = []
    current_attendance = {}
    selected_subject_obj = None
    
    if subject and date:
        try:
            subject_qs = Subject.objects.filter(id=subject)
            if semester:
                subject_qs = subject_qs.filter(semester_id=semester)
            subject_obj = subject_qs.first()
            if not subject_obj:
                raise ValueError('Invalid subject selection')

            if not _is_mentor_for_scope(request.user, subject_obj.semester, selected_divisions):
                mentor_error = 'You can view attendance only for classes where you are assigned as mentor.'
                raise ValueError('Mentor scope restricted')

            students = _get_attendance_students(subject_obj, selected_divisions)
            selected_subject_obj = subject_obj
            
            # Get current attendance records
            records = Attendance.objects.filter(subject=subject_obj, date=date)
            if selected_divisions:
                records = records.filter(student__division__in=selected_divisions)
            current_attendance = {record.student_id: record.status for record in records}
        except:
            pass
    
    context = {
        'form': form,
        'students': students,
        'current_attendance': current_attendance,
        'date': date,
        'semester': semester,
        'subject': subject,
        'selected_subject_obj': selected_subject_obj,
        'selected_divisions': selected_divisions,
        'mentor_error': mentor_error,
    }
    return render(request, 'attendance/mark_attendance.html', context)


@login_required(login_url='login')
def marks_entry(request):
    """Faculty/Admin marks entry for Mid/Final exams with attempt tracking."""
    faculty = getattr(request.user, 'faculty', None)
    if not request.user.is_staff and not faculty:
        return redirect('dashboard')

    selected_semester_id = request.GET.get('semester') or request.POST.get('semester')
    selected_divisions = [d for d in request.GET.getlist('division') if d]
    if request.method == 'POST':
        selected_divisions = [d for d in request.POST.getlist('division') if d]

    message = None
    students = []
    existing_marks = {}
    freeze_rule = None
    is_frozen = False
    mentor_error = None

    form_data = request.POST if request.method == 'POST' else request.GET
    form = FacultyMarksEntryForm(form_data or None, selected_semester_id=selected_semester_id)

    subject_obj = None
    exam_type = None
    exam_session = None
    attempt_no = None

    if form.is_valid():
        semester_obj = form.cleaned_data['semester']
        subject_obj = form.cleaned_data['subject']
        selected_divisions = form.cleaned_data.get('division') or []
        exam_type = form.cleaned_data['exam_type']
        exam_session = (form.cleaned_data['exam_session'] or '').strip().upper()
        attempt_no = form.cleaned_data['attempt_no']
        max_marks = form.cleaned_data['max_marks']
        pass_marks = form.cleaned_data.get('pass_marks')

        if subject_obj.semester_id != semester_obj.id:
            form.add_error('subject', 'Selected subject does not belong to selected semester.')
        else:
            if not _is_mentor_for_scope(request.user, semester_obj, selected_divisions):
                mentor_error = 'You can enter marks only for classes where you are assigned as mentor.'
            else:
                freeze_rule = MarksFreezeRule.objects.filter(
                    semester=semester_obj,
                    subject=subject_obj,
                    exam_type=exam_type,
                    exam_session=exam_session,
                    attempt_no=attempt_no,
                ).first()
                is_frozen = bool(freeze_rule and freeze_rule.is_frozen)

                students_qs = Student.objects.filter(branch=subject_obj.branch, semester=semester_obj)
                if selected_divisions:
                    students_qs = students_qs.filter(division__in=selected_divisions)
                students = list(students_qs.order_by('enrollment_no'))

                if request.method == 'POST':
                    action = (request.POST.get('action') or '').strip()

                if action == 'freeze_marks' and request.user.is_staff:
                    note = (request.POST.get('freeze_note') or '').strip()
                    freeze_rule, _ = MarksFreezeRule.objects.update_or_create(
                        semester=semester_obj,
                        subject=subject_obj,
                        exam_type=exam_type,
                        exam_session=exam_session,
                        attempt_no=attempt_no,
                        defaults={
                            'is_frozen': True,
                            'note': note,
                            'frozen_by': request.user,
                        },
                    )
                    is_frozen = True
                    message = 'Marks have been frozen for this exam combination.'

                elif action == 'unfreeze_marks' and request.user.is_staff:
                    note = (request.POST.get('freeze_note') or '').strip()
                    freeze_rule, _ = MarksFreezeRule.objects.update_or_create(
                        semester=semester_obj,
                        subject=subject_obj,
                        exam_type=exam_type,
                        exam_session=exam_session,
                        attempt_no=attempt_no,
                        defaults={
                            'is_frozen': False,
                            'note': note,
                            'frozen_by': request.user,
                        },
                    )
                    is_frozen = False
                    message = 'Marks have been reopened for editing.'

                elif action == 'save_marks':
                    if is_frozen:
                        message = 'Marks are frozen for this combination. Unfreeze first to edit.'
                    else:
                        audit_reason = (request.POST.get('audit_reason') or '').strip()
                        saved_count = 0
                        for student in students:
                            marks_key = f'marks_{student.id}'
                            absent_key = f'absent_{student.id}'
                            marks_text = (request.POST.get(marks_key) or '').strip()
                            is_absent_mark = request.POST.get(absent_key) == 'on'

                            if not is_absent_mark and not marks_text:
                                continue

                            marks_value = None
                            if not is_absent_mark:
                                try:
                                    marks_value = Decimal(marks_text)
                                except Exception:
                                    continue
                                if marks_value < 0 or marks_value > max_marks:
                                    continue

                            record = StudentMark.objects.filter(
                                student=student,
                                subject=subject_obj,
                                exam_type=exam_type,
                                exam_session=exam_session,
                                attempt_no=attempt_no,
                            ).first()

                            old_state = None
                            if record:
                                old_state = {
                                    'marks_obtained': record.marks_obtained,
                                    'is_absent': record.is_absent,
                                    'max_marks': record.max_marks,
                                    'pass_marks': record.pass_marks,
                                }
                                record.semester = semester_obj
                                record.max_marks = max_marks
                                record.pass_marks = pass_marks
                                record.marks_obtained = None if is_absent_mark else marks_value
                                record.is_absent = is_absent_mark
                                record.entered_by = faculty
                                record.save()
                            else:
                                record = StudentMark.objects.create(
                                    student=student,
                                    subject=subject_obj,
                                    semester=semester_obj,
                                    exam_type=exam_type,
                                    exam_session=exam_session,
                                    attempt_no=attempt_no,
                                    max_marks=max_marks,
                                    pass_marks=pass_marks,
                                    marks_obtained=None if is_absent_mark else marks_value,
                                    is_absent=is_absent_mark,
                                    entered_by=faculty,
                                )

                            changed = not old_state or (
                                old_state['marks_obtained'] != record.marks_obtained or
                                old_state['is_absent'] != record.is_absent or
                                old_state['max_marks'] != record.max_marks or
                                old_state['pass_marks'] != record.pass_marks
                            )

                            if changed and old_state:
                                MarksAuditTrail.objects.create(
                                    student_mark=record,
                                    student=student,
                                    subject=subject_obj,
                                    semester=semester_obj,
                                    exam_type=exam_type,
                                    exam_session=exam_session,
                                    attempt_no=attempt_no,
                                    old_marks=old_state['marks_obtained'],
                                    new_marks=record.marks_obtained,
                                    old_absent=old_state['is_absent'],
                                    new_absent=record.is_absent,
                                    old_max_marks=old_state['max_marks'],
                                    new_max_marks=record.max_marks,
                                    old_pass_marks=old_state['pass_marks'],
                                    new_pass_marks=record.pass_marks,
                                    reason=audit_reason,
                                    action='UPDATE',
                                    changed_by=request.user,
                                )

                            saved_count += 1
                        message = f'Marks saved for {saved_count} students.'

                marks_qs = StudentMark.objects.filter(
                    subject=subject_obj,
                    exam_type=exam_type,
                    exam_session=exam_session,
                    attempt_no=attempt_no,
                )
                existing_marks = {m.student_id: m for m in marks_qs}

    return render(request, 'results/marks_entry.html', {
        'form': form,
        'students': students,
        'existing_marks': existing_marks,
        'selected_divisions': selected_divisions,
        'message': message,
        'subject_obj': subject_obj,
        'exam_type': exam_type,
        'exam_session': exam_session,
        'attempt_no': attempt_no,
        'freeze_rule': freeze_rule,
        'is_frozen': is_frozen,
        'mentor_error': mentor_error,
    })


@login_required(login_url='login')
def manual_attendance_sheet_preview(request):
    """Preview manual attendance sheet with 31 day columns for print/PDF."""
    try:
        faculty = request.user.faculty
    except Exception:
        return redirect('dashboard')

    subject_id = (request.GET.get('subject') or '').strip()
    date_text = (request.GET.get('date') or '').strip()
    selected_divisions = [d for d in request.GET.getlist('division') if d]
    if not selected_divisions:
        single_division = request.GET.get('division')
        if single_division:
            selected_divisions = [single_division]

    if not subject_id:
        return redirect('mark_attendance')

    try:
        rows_per_side = int((request.GET.get('rows_per_page') or '40').strip())
    except Exception:
        rows_per_side = 40
    rows_per_side = max(20, min(80, rows_per_side))

    min_rows = 20
    max_rows = 80
    scale = (rows_per_side - min_rows) / float(max_rows - min_rows)
    row_height_screen_px = int(round(24 - (12 * scale)))  # 24px at 20 rows -> 12px at 80 rows
    row_height_print_px = int(round(18 - (8 * scale)))    # 18px at 20 rows -> 10px at 80 rows
    cell_pad_y_screen_px = max(1, int(round(2.5 - (1.5 * scale))))
    cell_pad_y_print_px = max(0, int(round(2.0 - (1.0 * scale))))

    subject_obj = Subject.objects.filter(id=subject_id).first()
    if not subject_obj:
        return redirect('mark_attendance')

    students_qs = Student.objects.filter(
        branch=subject_obj.branch,
        semester=subject_obj.semester,
    )
    if selected_divisions:
        students_qs = students_qs.filter(division__in=selected_divisions)

    students = list(students_qs.order_by('division', 'enrollment_no'))

    students_per_sheet = rows_per_side * 2
    division_order = selected_divisions[:] if selected_divisions else sorted({s.division for s in students if s.division})
    if not division_order:
        division_order = ['-']

    by_division = {}
    for student in students:
        key = student.division or '-'
        by_division.setdefault(key, []).append(student)

    def _build_side_rows(side_students, start_sr):
        rows = []
        for offset, student in enumerate(side_students):
            rows.append({
                'sr': start_sr + offset,
                'is_blank': False,
                'student': student,
            })
        while len(rows) < rows_per_side:
            rows.append({'sr': '', 'is_blank': True, 'student': None})
        return rows

    sheet_pairs = []
    for division_key in division_order:
        division_students = by_division.get(division_key, [])
        chunks = [
            division_students[i:i + students_per_sheet]
            for i in range(0, len(division_students), students_per_sheet)
        ]
        if not chunks:
            chunks = [[]]

        total_pairs = len(chunks)
        for pair_index, chunk in enumerate(chunks, start=1):
            front_students = chunk[:rows_per_side]
            back_students = chunk[rows_per_side:students_per_sheet]

            sheet_pairs.append({
                'division': division_key,
                'pair_no': pair_index,
                'total_pairs': total_pairs,
                'front_rows': _build_side_rows(front_students, 1),
                'back_rows': _build_side_rows(back_students, rows_per_side + 1),
                'has_back_rows': bool(back_students),
            })

    month_name = ''
    year_value = ''
    month_number = None
    formatted_date = ''
    if date_text:
        try:
            selected_date = datetime.strptime(date_text, '%Y-%m-%d').date()
            month_name = selected_date.strftime('%B')
            year_value = selected_date.strftime('%Y')
            month_number = selected_date.month
            formatted_date = selected_date.strftime('%d-%m-%Y')
        except Exception:
            formatted_date = date_text

    if not month_name or not year_value:
        today = timezone.localdate()
        month_name = today.strftime('%B')
        year_value = today.strftime('%Y')
        month_number = today.month

    try:
        year_int = int(year_value)
    except Exception:
        year_int = timezone.localdate().year
    if not month_number:
        month_number = timezone.localdate().month
    days_in_month = calendar.monthrange(year_int, month_number)[1]

    context = {
        'subject_obj': subject_obj,
        'students': students,
        'sheet_pairs': sheet_pairs,
        'rows_per_side': rows_per_side,
        'students_per_sheet': students_per_sheet,
        'row_height_screen_px': row_height_screen_px,
        'row_height_print_px': row_height_print_px,
        'cell_pad_y_screen_px': cell_pad_y_screen_px,
        'cell_pad_y_print_px': cell_pad_y_print_px,
        'selected_divisions': selected_divisions,
        'date_text': date_text,
        'formatted_date': formatted_date,
        'month_name': month_name,
        'year_value': year_value,
        'generated_on': timezone.localtime().strftime('%d-%m-%Y %I:%M %p'),
        'day_columns': list(range(1, days_in_month + 1)),
        'institute_line1': 'KADI SARVA VISHWAVIDYALAYA',
        'institute_line2': 'LDRP Institute of Technology and Research Gandhinagar',
    }
    return render(request, 'attendance/manual_attendance_sheet_preview.html', context)

# ============= MARKS VIEWS REMOVED =============

# ============= EXAM SEATING (EXISTING) =============

def parse_rooms(room_text):
    rooms = []
    seen = set()
    for part in room_text.split(','):
        part = part.strip()
        if '-' in part:
            start, end = part.split('-')
            for r in range(int(start), int(end) + 1):
                room = str(r)
                if room not in seen:
                    rooms.append(room)
                    seen.add(room)
        else:
            if part and part not in seen:
                rooms.append(part)
                seen.add(part)
    return rooms

@login_required(login_url='login')
def allocate(request):
    """Exam seating allocation (Admin only)"""
    # Allow faculty to generate seating as requested
    
    if request.method == "POST":
        form = CEForm(request.POST, request.FILES)
        if form.is_valid():
            CESeating.objects.all().delete()
            
            semester = form.cleaned_data['semester']
            month = form.cleaned_data['month']
            year = form.cleaned_data['year']
            date_from = form.cleaned_data['date_from']
            date_to = form.cleaned_data['date_to']
            examination_name = form.cleaned_data.get('examination_name', '')
            note = form.cleaned_data.get('note', '')
            start_exam_no = form.cleaned_data['start_exam_no']
            total = form.cleaned_data['total_students']
            preferred = form.cleaned_data['preferred_per_room']
            max_limit = form.cleaned_data['max_per_room']
            
            rooms = parse_rooms(form.cleaned_data['rooms'])
            room_count = len(rooms)
            
            # Capacity Check
            if total > room_count * max_limit:
                return HttpResponse("ERROR: Not enough rooms. Please add more rooms.", status=400)
            
            # Priority-based allocation:
            # 1. Try preferred amount per room first
            # 2. If can't fit all students, incrementally increase towards max
            # 3. Adjust smartly at end if few students remain
            
            # Start with preferred amount
            per_room = preferred
            rooms_needed = (total + per_room - 1) // per_room  # ceiling
            
            # If can't fit with preferred, step up towards max
            while rooms_needed > room_count and per_room < max_limit:
                per_room += 1
                rooms_needed = (total + per_room - 1) // per_room
            
            # Final check
            if rooms_needed > room_count:
                return HttpResponse("ERROR: Not enough rooms. Please add more rooms.", status=400)
            
            # Allocate rooms using calculated per_room amount
            room_allocations = []
            remaining = total
            
            for i in range(rooms_needed):
                if remaining <= 0:
                    break
                
                if remaining >= per_room:
                    # Fill with current per_room amount
                    room_allocations.append(per_room)
                    remaining -= per_room
                else:
                    # Last room with remaining students
                    if remaining > 0:
                        # If very few left (less than half), redistribute with last room
                        if remaining < per_room // 2 and len(room_allocations) > 0:
                            last_room = room_allocations.pop()
                            total_for_last_two = remaining + last_room
                            mid = total_for_last_two // 2
                            room_allocations.append(mid)
                            room_allocations.append(total_for_last_two - mid)
                        else:
                            room_allocations.append(remaining)
                        remaining = 0
            
            # Pad unused rooms with 0 count
            while len(room_allocations) < room_count:
                room_allocations.append(0)
            
            # Create seating records (skip empty rooms)
            seat_no = start_exam_no
            for i, room in enumerate(rooms):
                count = room_allocations[i]
                if count > 0:  # Only create records for non-empty rooms
                    start = seat_no
                    end = seat_no + count - 1
                    
                    CESeating.objects.create(
                        room_no=room,
                        seat_from=start,
                        seat_to=end,
                        count=count
                    )
                    seat_no = end + 1
            
            from urllib.parse import urlencode
            params = urlencode({
                'semester': semester,
                'month': month,
                'year': year,
                'date_from': date_from,
                'date_to': date_to,
                'examination_name': examination_name,
                'note': note
            })
            return redirect(f"/preview/?{params}")
    else:
        form = CEForm()
    
    return render(request, "allocate.html", {"form": form})

# ============= ADMIN HELPERS =============

@login_required(login_url='login')
def manage_roles(request):
    """Admin page to create new faculty users and assign faculty roles."""
    if not request.user.is_staff:
        return redirect('dashboard')

    # Users without a faculty profile
    users_without_faculty = User.objects.filter(faculty__isnull=True).order_by('username')
    subjects = Subject.objects.all().order_by('code')
    all_faculty = Faculty.objects.select_related('user').order_by('user__username')
    semesters = Semester.objects.all().order_by('number')
    division_choices = Student.DIVISION_CHOICES
    mentor_assignments = MentorAssignment.objects.select_related('faculty__user', 'semester').order_by('semester__number', 'division')

    message = None
    generated_password = None

    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_faculty':
            # Create new faculty user
            full_name = request.POST.get('full_name', '').strip()
            email = request.POST.get('email', '').strip()
            employee_id = request.POST.get('employee_id', '').strip()
            subject_ids = request.POST.getlist('subjects')
            mentor_semester_id = request.POST.get('mentor_semester')
            mentor_division = (request.POST.get('mentor_division') or '').strip().upper()
            
            if not full_name or not employee_id or not email:
                message = "Full name, email, and employee ID are required."
            else:
                username = email

                if User.objects.filter(username=username).exists():
                    message = "Email is already in use as a username."
                else:
                    # Generate random password
                    import random
                    import string
                    password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
                    generated_password = password
                    
                    # Create user
                    new_user = User.objects.create_user(
                        username=username,
                        email=email,
                        password=password,
                        first_name=full_name.split()[0] if full_name else '',
                        last_name=' '.join(full_name.split()[1:]) if len(full_name.split()) > 1 else ''
                    )
                    
                    # Create faculty profile
                    faculty = Faculty.objects.create(
                        user=new_user,
                        employee_id=employee_id,
                        must_change_password=True,
                    )
                    
                    if subject_ids:
                        faculty.subjects.set(Subject.objects.filter(id__in=subject_ids))

                    if mentor_semester_id and mentor_division:
                        mentor_semester = Semester.objects.filter(id=mentor_semester_id).first()
                        if mentor_semester and mentor_division in dict(Student.DIVISION_CHOICES):
                            existing_qs = MentorAssignment.objects.filter(
                                semester=mentor_semester,
                                division=mentor_division,
                            )
                            if existing_qs.count() >= 2 and not existing_qs.filter(faculty=faculty).exists():
                                message = f"Faculty created. Could not add mentor: S{mentor_semester.number}-{mentor_division} already has two mentors."
                            else:
                                MentorAssignment.objects.get_or_create(
                                    faculty=faculty,
                                    semester=mentor_semester,
                                    division=mentor_division,
                                )
                    
                    message = f"Faculty created successfully! Username: {username}, Password: {password}"
        
        elif action == 'assign':
            user_id = request.POST.get('user_id')
            employee_id = request.POST.get('employee_id')
            subject_ids = request.POST.getlist('subjects')
            mentor_semester_id = request.POST.get('mentor_semester')
            mentor_division = (request.POST.get('mentor_division') or '').strip().upper()
            try:
                target_user = User.objects.get(id=user_id)
                faculty, created = Faculty.objects.get_or_create(user=target_user, defaults={
                    'employee_id': employee_id or f'EMP{target_user.id:04d}'
                })
                if not created and employee_id:
                    faculty.employee_id = employee_id
                    faculty.save()
                if subject_ids:
                    faculty.subjects.set(Subject.objects.filter(id__in=subject_ids))

                if mentor_semester_id and mentor_division:
                    mentor_semester = Semester.objects.filter(id=mentor_semester_id).first()
                    if mentor_semester and mentor_division in dict(Student.DIVISION_CHOICES):
                        existing_qs = MentorAssignment.objects.filter(
                            semester=mentor_semester,
                            division=mentor_division,
                        )
                        if existing_qs.count() >= 2 and not existing_qs.filter(faculty=faculty).exists():
                            message = f"Assigned faculty role, but S{mentor_semester.number}-{mentor_division} already has two mentors."
                        else:
                            MentorAssignment.objects.get_or_create(
                                faculty=faculty,
                                semester=mentor_semester,
                                division=mentor_division,
                            )
                if not message:
                    message = f"Assigned faculty role to {target_user.username}."
            except User.DoesNotExist:
                message = "Selected user does not exist."
        
        elif action == 'seed_demo':
            # Create a demo faculty user and link a faculty profile
            demo_username = 'faculty1'
            demo_password = 'faculty123'
            demo_email = 'faculty1@example.com'
            target_user, created = User.objects.get_or_create(username=demo_username, defaults={
                'email': demo_email
            })
            if created:
                target_user.set_password(demo_password)
                target_user.save()
            faculty, fac_created = Faculty.objects.get_or_create(user=target_user, defaults={
                'employee_id': 'EMP0001'
            })
            if subjects.exists():
                # Assign first subject by default to make dashboards work
                faculty.subjects.add(subjects.first())
            message = "Demo faculty user created: username 'faculty1', password 'faculty123'."

        elif action == 'update_faculty':
            faculty_id = request.POST.get('faculty_id')
            full_name = request.POST.get('full_name', '').strip()
            email = request.POST.get('email', '').strip()
            employee_id = request.POST.get('employee_id', '').strip()
            new_password = request.POST.get('new_password', '')
            subject_ids = request.POST.getlist('subjects')
            mentor_semester_id = request.POST.get('mentor_semester')
            mentor_division = (request.POST.get('mentor_division') or '').strip().upper()

            try:
                faculty = Faculty.objects.select_related('user').get(id=faculty_id)
            except Faculty.DoesNotExist:
                faculty = None

            if not faculty:
                message = 'Selected faculty does not exist.'
            else:
                user = faculty.user

                if email:
                    if User.objects.filter(username=email).exclude(id=user.id).exists():
                        message = 'Email is already in use as a username.'
                    else:
                        user.username = email
                        user.email = email

                if full_name:
                    parts = full_name.split()
                    user.first_name = parts[0]
                    user.last_name = ' '.join(parts[1:]) if len(parts) > 1 else ''

                if new_password:
                    user.set_password(new_password)
                    faculty.must_change_password = True

                if employee_id:
                    faculty.employee_id = employee_id

                user.save()
                faculty.save()

                if subject_ids:
                    faculty.subjects.set(Subject.objects.filter(id__in=subject_ids))
                else:
                    faculty.subjects.clear()

                if mentor_semester_id and mentor_division:
                    mentor_semester = Semester.objects.filter(id=mentor_semester_id).first()
                    if mentor_semester and mentor_division in dict(Student.DIVISION_CHOICES):
                        existing_qs = MentorAssignment.objects.filter(
                            semester=mentor_semester,
                            division=mentor_division,
                        )
                        if existing_qs.count() >= 2 and not existing_qs.filter(faculty=faculty).exists():
                            message = f"Updated faculty, but S{mentor_semester.number}-{mentor_division} already has two mentors."
                        else:
                            MentorAssignment.objects.get_or_create(
                                faculty=faculty,
                                semester=mentor_semester,
                                division=mentor_division,
                            )

                if not message:
                    message = f"Updated faculty {user.username}."

        elif action == 'remove_mentor_assignment':
            assignment_id = request.POST.get('assignment_id')
            deleted, _ = MentorAssignment.objects.filter(id=assignment_id).delete()
            if deleted:
                message = 'Mentor assignment removed.'
            else:
                message = 'Mentor assignment not found.'

        elif action == 'delete_faculty':
            faculty_id = request.POST.get('faculty_id')
            try:
                faculty = Faculty.objects.select_related('user').get(id=faculty_id)
            except Faculty.DoesNotExist:
                faculty = None

            if not faculty:
                message = 'Selected faculty does not exist.'
            elif faculty.user_id == request.user.id:
                message = 'You cannot delete your own account.'
            else:
                username = faculty.user.username
                faculty.user.delete()
                message = f"Deleted faculty user {username}."

        # Refresh lists after changes
        users_without_faculty = User.objects.filter(faculty__isnull=True).order_by('username')
        all_faculty = Faculty.objects.select_related('user').order_by('user__username')
        mentor_assignments = MentorAssignment.objects.select_related('faculty__user', 'semester').order_by('semester__number', 'division')

    context = {
        'users_without_faculty': users_without_faculty,
        'subjects': subjects,
        'all_faculty': all_faculty,
        'semesters': semesters,
        'division_choices': division_choices,
        'mentor_assignments': mentor_assignments,
        'message': message,
        'generated_password': generated_password,
    }
    return render(request, 'admin/manage_roles.html', context)

@login_required(login_url='login')
def assign_subjects(request):
    """Admin page to assign subjects to faculty users with filters."""
    if not request.user.is_staff:
        return redirect('dashboard')

    faculties = Faculty.objects.select_related('user').order_by('user__username')
    branches = Branch.objects.all().order_by('name')
    semesters = Semester.objects.all().order_by('number')

    faculty_id = request.GET.get('faculty') or request.POST.get('faculty')
    branch_id = request.GET.get('branch') or request.POST.get('branch')
    semester_id = request.GET.get('semester') or request.POST.get('semester')

    selected_faculty = None
    subjects = Subject.objects.none()
    assigned_ids = set()
    message = None

    if faculty_id:
        selected_faculty = get_object_or_404(Faculty, id=faculty_id)

        # Filter subjects optionally by branch and semester
        subj_qs = Subject.objects.all()
        if branch_id:
            subj_qs = subj_qs.filter(branch_id=branch_id)
        if semester_id:
            subj_qs = subj_qs.filter(semester_id=semester_id)
        subjects = subj_qs.order_by('code')

        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'save_assignments':
                selected_subject_ids = [int(sid) for sid in request.POST.getlist('subjects')]
                # Only consider subjects visible in current filter for update set/add/remove
                visible_ids = list(subjects.values_list('id', flat=True))
                current_set = set(selected_faculty.subjects.values_list('id', flat=True))

                # Add newly selected among visible
                to_add = set(selected_subject_ids) - current_set
                if to_add:
                    selected_faculty.subjects.add(*Subject.objects.filter(id__in=to_add))

                # Remove deselected among visible
                to_remove = (current_set & set(visible_ids)) - set(selected_subject_ids)
                if to_remove:
                    selected_faculty.subjects.remove(*Subject.objects.filter(id__in=to_remove))

                message = 'Assignments updated.'

        assigned_ids = set(selected_faculty.subjects.values_list('id', flat=True))

    context = {
        'faculties': faculties,
        'branches': branches,
        'semesters': semesters,
        'selected_faculty': selected_faculty,
        'subjects': subjects,
        'assigned_ids': assigned_ids,
        'branch_id': branch_id,
        'semester_id': semester_id,
        'message': message,
    }
    return render(request, 'admin/assign_subjects.html', context)

# ============= SUBJECTS BY SEMESTER =============

@login_required(login_url='login')
def manage_subjects(request):
    """Admin page to list subjects by branch/semester and add new ones."""
    if not request.user.is_staff:
        return redirect('dashboard')

    selected_branch_id = request.GET.get('branch')
    selected_semester_id = request.GET.get('semester')
    if request.method == 'POST':
        selected_branch_id = request.POST.get('branch') or selected_branch_id
        selected_semester_id = request.POST.get('semester') or selected_semester_id

    branches = Branch.objects.all().order_by('name')
    semesters = Semester.objects.all().order_by('number')

    subjects = Subject.objects.none()
    if selected_branch_id and selected_semester_id:
        subjects = Subject.objects.filter(branch_id=selected_branch_id, semester_id=selected_semester_id).order_by('code')

    message = None
    errors = []
    row_count = 10

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'bulk_add':
            branch_id = request.POST.get('branch')
            semester_id = request.POST.get('semester')
            if not branch_id or not semester_id:
                message = "Please select branch and semester first."
            else:
                branch = get_object_or_404(Branch, id=branch_id)
                semester = get_object_or_404(Semester, id=semester_id)
                try:
                    row_count = int(request.POST.get('row_count', row_count))
                except ValueError:
                    row_count = 10

                created = 0
                updated = 0

                for i in range(row_count):
                    code = (request.POST.get(f'code_{i}') or '').strip()
                    name = (request.POST.get(f'name_{i}') or '').strip()
                    credit_raw = (request.POST.get(f'credit_{i}') or '').strip()
                    is_elective = request.POST.get(f'elective_{i}') == 'on'
                    elective_group = (request.POST.get(f'group_{i}') or '').strip()

                    if not code and not name:
                        continue
                    if not code or not name:
                        errors.append(f"Row {i + 1}: Code and Name are required.")
                        continue

                    credit = Decimal('0')
                    if credit_raw:
                        try:
                            credit = Decimal(credit_raw)
                        except InvalidOperation:
                            errors.append(f"Row {i + 1}: Credit must be a number.")
                            continue
                        if credit < 0:
                            errors.append(f"Row {i + 1}: Credit must be 0 or greater.")
                            continue

                    try:
                        subject, created_flag = Subject.objects.update_or_create(
                            code=code,
                            defaults={
                                'name': name,
                                'branch': branch,
                                'semester': semester,
                                'credit': credit,
                                'is_elective': is_elective,
                                'elective_group': elective_group,
                            }
                        )
                        if created_flag:
                            created += 1
                        else:
                            updated += 1
                    except IntegrityError:
                        errors.append(f"Row {i + 1}: Could not save subject {code}.")

                if not errors:
                    message = f"Saved {created} subjects and updated {updated} subjects."
                else:
                    message = f"Saved {created} subjects and updated {updated} subjects with some errors."

                subjects = Subject.objects.filter(branch=branch, semester=semester).order_by('code')

        elif action == 'bulk_move':
            from_branch_id = request.POST.get('from_branch')
            from_semester_id = request.POST.get('from_semester')
            to_branch_id = request.POST.get('to_branch')
            to_semester_id = request.POST.get('to_semester')

            if not (from_branch_id and from_semester_id and to_branch_id and to_semester_id):
                message = "Please select both source and destination branch/semester."
            elif from_branch_id == to_branch_id and from_semester_id == to_semester_id:
                message = "Source and destination are the same. Nothing to move."
            else:
                from_branch = get_object_or_404(Branch, id=from_branch_id)
                from_semester = get_object_or_404(Semester, id=from_semester_id)
                to_branch = get_object_or_404(Branch, id=to_branch_id)
                to_semester = get_object_or_404(Semester, id=to_semester_id)

                moved = Subject.objects.filter(branch=from_branch, semester=from_semester).update(
                    branch=to_branch,
                    semester=to_semester
                )
                message = f"Moved {moved} subjects to {to_branch.name}, Semester {to_semester.number}."
                selected_branch_id = str(to_branch.id)
                selected_semester_id = str(to_semester.id)
                subjects = Subject.objects.filter(branch=to_branch, semester=to_semester).order_by('code')

    context = {
        'branches': branches,
        'semesters': semesters,
        'subjects': subjects,
        'selected_branch_id': selected_branch_id,
        'selected_semester_id': selected_semester_id,
        'message': message,
        'errors': errors,
        'row_count': row_count,
        'row_indexes': list(range(row_count)),
    }
    return render(request, 'subjects/manage_subjects.html', context)


@login_required(login_url='login')
def edit_subject(request, subject_id):
    """Edit a single subject (Admin only)."""
    if not request.user.is_staff:
        return redirect('dashboard')

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            subject = form.save()
            params = urlencode({'branch': subject.branch_id, 'semester': subject.semester_id})
            return redirect(f"/subjects/?{params}")
    else:
        form = SubjectForm(instance=subject)

    return render(request, 'subjects/edit_subject.html', {
        'form': form,
        'subject': subject,
    })


@login_required(login_url='login')
def delete_subject(request, subject_id):
    """Delete a single subject with confirmation (Admin only)."""
    if not request.user.is_staff:
        return redirect('dashboard')

    subject = get_object_or_404(Subject, id=subject_id)

    if request.method == 'POST':
        branch_id = subject.branch_id
        semester_id = subject.semester_id
        subject.delete()
        params = urlencode({'branch': branch_id, 'semester': semester_id})
        return redirect(f"/subjects/?{params}")

    return render(request, 'subjects/delete_subject.html', {
        'subject': subject,
    })


# ============= STUDENT UPLOAD =============

@login_required(login_url='login')
def upload_students(request):
    """Upload a sheet (CSV/XLSX) of students and bulk insert into a selected branch+semester."""
    if not request.user.is_staff:
        return redirect('dashboard')

    message = None
    results = None
    preview_rows = None

    if request.method == 'POST':
        form = StudentUploadForm(request.POST, request.FILES)
        if form.is_valid():
            branch = form.cleaned_data['branch']
            semester = form.cleaned_data['semester']
            f = form.cleaned_data['file']

            # Read file using pandas
            try:
                if str(f.name).lower().endswith('.csv'):
                    df = pd.read_csv(f)
                else:
                    df = pd.read_excel(f)
            except Exception as e:
                message = f"Failed to read file: {e}"
                df = None

            if df is not None:
                # Normalize column names
                df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

                # Map common variants
                colmap = {
                    'enrollment_no': ['enrollment_no', 'enrollment', 'enrollment_number', 'enrollmentno'],
                    'name': ['name', 'student_name', 'fullname'],
                    'division': ['division', 'div'],
                    'admission_year': ['admission_year', 'admissionyear', 'year_of_admission', 'year'],
                    'mentor_name': ['mentor_name', 'mentor', 'class_teacher'],
                    'email': ['email', 'mail'],
                    'phone': ['phone', 'mobile', 'contact'],
                    'elective_codes': ['elective', 'electives', 'elective_subjects', 'elective_codes', 'subjects', 'subject_codes', 'subject_code']
                }

                resolved = {}
                for target, variants in colmap.items():
                    for v in variants:
                        if v in df.columns:
                            resolved[target] = v
                            break

                required_missing = [k for k in ['enrollment_no', 'name'] if k not in resolved]
                if required_missing:
                    message = f"Missing required columns: {', '.join(required_missing)}."
                else:
                    # Prepare rows
                    rows = []
                    errors = []
                    elective_lookup = {
                        s.code.strip().upper(): s
                        for s in Subject.objects.filter(branch=branch, semester=semester, is_elective=True)
                    }
                    subject_lookup = {
                        s.code.strip().upper(): s
                        for s in Subject.objects.filter(branch=branch, semester=semester)
                    }
                    for i, row in df.iterrows():
                        enroll = str(row[resolved['enrollment_no']]).strip()
                        name = str(row[resolved['name']]).strip()
                        division = str(row[resolved.get('division', '')]).strip() if resolved.get('division') else ''
                        admission_year = None
                        if resolved.get('admission_year'):
                            year_value = row.get(resolved['admission_year'])
                            if not pd.isna(year_value):
                                try:
                                    admission_year = int(float(str(year_value).strip()))
                                except Exception:
                                    errors.append(f"Row {i+1}: invalid admission year")
                                    continue
                                if admission_year < 2000 or admission_year > 2100:
                                    errors.append(f"Row {i+1}: admission year must be between 2000 and 2100")
                                    continue
                        mentor_name = str(row[resolved.get('mentor_name', '')]).strip() if resolved.get('mentor_name') else ''
                        email = str(row[resolved.get('email', '')]) if resolved.get('email') else ''
                        phone = str(row[resolved.get('phone', '')]) if resolved.get('phone') else ''
                        elective_raw = ''
                        if resolved.get('elective_codes'):
                            elective_value = row.get(resolved['elective_codes'])
                            elective_raw = '' if pd.isna(elective_value) else str(elective_value).strip()
                        elective_codes = []
                        if elective_raw:
                            elective_codes = [
                                code.strip().upper()
                                for code in re.split(r'[;,|]+', elective_raw)
                                if code and code.strip()
                            ]

                        if not enroll or not name:
                            errors.append(f"Row {i+1}: missing required fields")
                            continue

                        rows.append({
                            'row_index': i + 1,
                            'enrollment_no': enroll,
                            'name': name,
                            'division': division or 'A',
                            'admission_year': admission_year,
                            'mentor_name': mentor_name,
                            'email': email,
                            'phone': phone,
                            'elective_codes': elective_codes,
                            'elective_display': elective_raw,
                        })

                    # Insert records (skip duplicates)
                    created = 0
                    skipped = 0
                    for r in rows:
                        if Student.objects.filter(enrollment_no=r['enrollment_no']).exists():
                            skipped += 1
                            continue
                        student = Student.objects.create(
                            enrollment_no=r['enrollment_no'],
                            name=r['name'],
                            division=r['division'],
                            admission_year=r['admission_year'],
                            mentor_name=r['mentor_name'],
                            email=r['email'],
                            phone=r['phone'],
                            branch=branch,
                            semester=semester,
                        )
                        if r['elective_codes']:
                            for code in r['elective_codes']:
                                subject = elective_lookup.get(code)
                                if subject is None:
                                    if code in subject_lookup:
                                        errors.append(
                                            f"Row {r['row_index']}: {code} is not marked as an elective subject."
                                        )
                                    else:
                                        errors.append(
                                            f"Row {r['row_index']}: elective code {code} not found in this branch/semester."
                                        )
                                    continue
                                StudentSubject.objects.get_or_create(student=student, subject=subject)
                        created += 1

                    results = {
                        'created': created,
                        'skipped': skipped,
                        'errors': errors,
                        'total': len(rows)
                    }
                    preview_rows = rows[:20]
                    message = f"Upload complete: {created} created, {skipped} skipped."
        else:
            message = "Please select branch, semester and a valid file."
    else:
        form = StudentUploadForm()

    return render(request, 'students/upload_students.html', {
        'form': form,
        'message': message,
        'results': results,
        'preview_rows': preview_rows,
    })


# ============= MARKSHEET UPLOAD =============

@login_required(login_url='login')
def upload_marksheet(request):
    """Admin upload of marks via CSV/XLSX."""
    if not request.user.is_staff:
        return redirect('dashboard')

    message = None
    results = None
    preview_rows = None

    if request.method == 'POST':
        form = MarksheetUploadForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data['file']
            is_csv = str(f.name).lower().endswith('.csv')

            try:
                if is_csv:
                    df = pd.read_csv(f)
                else:
                    f.seek(0)
                    df = pd.read_excel(f)
            except Exception as e:
                message = f"Failed to read file: {e}"
                df = None

            if df is not None:
                df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]

                colmap = {
                    'enrollment_no': ['enrollment_no', 'enrollment', 'enrollment_number', 'enrollmentno'],
                    'semester': ['semester', 'sem', 'sem_no'],
                    'exam_session': ['exam_session', 'exam', 'session', 'exam_month'],
                    'issued_date': ['issued_date', 'issue_date', 'date'],
                    'spi': ['spi'],
                    'cpi': ['cpi'],
                    'earned_credits': ['earned_credits', 'earned_credit'],
                    'earned_grade_points': ['earned_grade_points', 'earned_gp', 'earned_points'],
                    'total_credits': ['total_credits', 'total_credit'],
                    'total_grade_points': ['total_grade_points', 'total_gp', 'total_points'],
                    'result_status': ['result_status', 'result'],
                    'course_code': ['course_code', 'code', 'subject_code', 'coursecode'],
                    'course_name': ['course_name', 'subject_name', 'name'],
                    'course_credit': ['course_credit', 'credit', 'course_credits'],
                    'grade': ['grade'],
                }

                resolved = {}
                for target, variants in colmap.items():
                    for v in variants:
                        if v in df.columns:
                            resolved[target] = v
                            break

                required = ['enrollment_no', 'semester', 'exam_session', 'course_code', 'course_name', 'course_credit', 'grade']
                required_missing = [k for k in required if k not in resolved]

                if required_missing and not is_csv:
                    try:
                        f.seek(0)
                        raw_df = pd.read_excel(f, header=None)
                        parsed_rows, parse_error = _extract_consolidated_marks_rows(raw_df)
                    except Exception as exc:
                        parsed_rows, parse_error = None, f'Failed to parse consolidated sheet: {exc}'

                    if parsed_rows:
                        df = pd.DataFrame(parsed_rows)
                        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
                        resolved = {k: k for k in colmap.keys() if k in df.columns}
                        required_missing = [k for k in required if k not in resolved]
                    elif parse_error:
                        message = parse_error

                if required_missing:
                    if not message:
                        message = f"Missing required columns: {', '.join(required_missing)}."
                else:
                    errors = []
                    preview_rows = []
                    sheet_map = {}
                    entries_map = {}

                    for i, row in df.iterrows():
                        enroll = str(row[resolved['enrollment_no']]).strip()
                        sem_raw = str(row[resolved['semester']]).strip()
                        exam_session = str(row[resolved['exam_session']]).strip()
                        course_code = str(row[resolved['course_code']]).strip()
                        course_name = str(row[resolved['course_name']]).strip()

                        if not enroll or not sem_raw or not exam_session or not course_code or not course_name:
                            errors.append(f"Row {i+1}: missing required fields")
                            continue

                        try:
                            semester_no = int(float(sem_raw))
                        except Exception:
                            errors.append(f"Row {i+1}: invalid semester for {enroll}")
                            continue

                        try:
                            course_credit = float(row[resolved['course_credit']])
                        except Exception:
                            errors.append(f"Row {i+1}: invalid course credit for {course_code}")
                            continue

                        student = Student.objects.filter(enrollment_no__iexact=enroll).select_related('semester').first()
                        if not student:
                            errors.append(f"Row {i+1}: student not found ({enroll})")
                            continue

                        semester_obj = Semester.objects.filter(number=semester_no).first()
                        if not semester_obj:
                            errors.append(f"Row {i+1}: semester not found ({semester_no})")
                            continue

                        if student.semester_id != semester_obj.id:
                            errors.append(f"Row {i+1}: semester mismatch for {enroll}")
                            continue

                        issued_date_val = None
                        if resolved.get('issued_date') and not pd.isna(row[resolved['issued_date']]):
                            try:
                                issued_date_val = pd.to_datetime(row[resolved['issued_date']]).date()
                            except Exception:
                                errors.append(f"Row {i+1}: invalid issued date for {enroll}")
                                continue

                        spi = row[resolved['spi']] if resolved.get('spi') and not pd.isna(row[resolved['spi']]) else None
                        cpi = row[resolved['cpi']] if resolved.get('cpi') and not pd.isna(row[resolved['cpi']]) else None
                        earned_credits = row[resolved['earned_credits']] if resolved.get('earned_credits') and not pd.isna(row[resolved['earned_credits']]) else None
                        earned_grade_points = row[resolved['earned_grade_points']] if resolved.get('earned_grade_points') and not pd.isna(row[resolved['earned_grade_points']]) else None
                        total_credits = row[resolved['total_credits']] if resolved.get('total_credits') and not pd.isna(row[resolved['total_credits']]) else None
                        total_grade_points = row[resolved['total_grade_points']] if resolved.get('total_grade_points') and not pd.isna(row[resolved['total_grade_points']]) else None
                        result_status = str(row[resolved['result_status']]).strip() if resolved.get('result_status') and not pd.isna(row[resolved['result_status']]) else 'PASS'

                        key = (student.id, semester_obj.id, exam_session)
                        if key not in sheet_map:
                            sheet_map[key] = {
                                'student': student,
                                'semester': semester_obj,
                                'exam_session': exam_session,
                                'issued_date': issued_date_val,
                                'spi': spi,
                                'cpi': cpi,
                                'earned_credits': earned_credits,
                                'earned_grade_points': earned_grade_points,
                                'total_credits': total_credits,
                                'total_grade_points': total_grade_points,
                                'result_status': result_status,
                            }

                        entries_map.setdefault(key, []).append({
                            'course_code': course_code,
                            'course_name': course_name,
                            'course_credit': course_credit,
                            'grade': str(row[resolved['grade']]).strip() if resolved.get('grade') else '',
                        })

                        if len(preview_rows) < 20:
                            preview_rows.append({
                                'enrollment_no': enroll,
                                'semester': semester_no,
                                'exam_session': exam_session,
                                'course_code': course_code,
                                'course_name': course_name,
                                'course_credit': course_credit,
                                'grade': str(row[resolved['grade']]).strip() if resolved.get('grade') else '',
                            })

                    created = 0
                    updated = 0
                    for key, sheet_data in sheet_map.items():
                        sheet, was_created = ResultSheet.objects.update_or_create(
                            student=sheet_data['student'],
                            semester=sheet_data['semester'],
                            exam_session=sheet_data['exam_session'],
                            defaults={
                                'issued_date': sheet_data['issued_date'],
                                'spi': sheet_data['spi'],
                                'cpi': sheet_data['cpi'],
                                'earned_credits': sheet_data['earned_credits'],
                                'earned_grade_points': sheet_data['earned_grade_points'],
                                'total_credits': sheet_data['total_credits'],
                                'total_grade_points': sheet_data['total_grade_points'],
                                'result_status': sheet_data['result_status'],
                            }
                        )
                        if was_created:
                            created += 1
                        else:
                            updated += 1

                        ResultEntry.objects.filter(result_sheet=sheet).delete()
                        entries = entries_map.get(key, [])
                        ResultEntry.objects.bulk_create([
                            ResultEntry(
                                result_sheet=sheet,
                                course_code=e['course_code'],
                                course_name=e['course_name'],
                                course_credit=e['course_credit'],
                                grade=e['grade']
                            ) for e in entries
                        ])

                    results = {
                        'created': created,
                        'updated': updated,
                        'errors': errors,
                        'total': len(df)
                    }
                    message = f"Upload complete: {created} created, {updated} updated."
                    if created or updated:
                        _create_result_announcement(
                            sheet_map=sheet_map,
                            created_count=created,
                            updated_count=updated,
                            user=request.user,
                        )
        else:
            message = "Please select a valid file."
    else:
        form = MarksheetUploadForm()

    return render(request, 'results/upload_marks.html', {
        'form': form,
        'message': message,
        'results': results,
        'preview_rows': preview_rows,
    })


# ============= PUBLIC RESULT LOOKUP =============

def public_result(request):
    form = ResultLookupForm(request.POST or None)
    context = {'form': form}

    if request.method == 'POST' and form.is_valid():
        enrollment = form.cleaned_data['enrollment_no'].strip()
        semester_no = form.cleaned_data['semester']
        student = Student.objects.filter(
            enrollment_no__iexact=enrollment,
            semester__number=semester_no
        ).select_related('branch', 'semester').first()

        if not student:
            context['error'] = 'Result not found for the given enrollment and semester.'
        else:
            result_context, error = _build_result_context(student, semester_no=semester_no)
            if error:
                context['error'] = error
            else:
                context.update(result_context)

    return render(request, 'results/result_lookup.html', context)

def public_result_pdf(request):
    enrollment = request.GET.get('enrollment_no', '').strip()
    semester = request.GET.get('semester', '').strip()

    try:
        semester_no = int(semester)
    except Exception:
        return HttpResponse('Invalid request.', status=400)

    student = Student.objects.filter(
        enrollment_no__iexact=enrollment,
        semester__number=semester_no
    ).select_related('branch', 'semester').first()

    if not student:
        return HttpResponse('Result not found.', status=404)

    result_context, error = _build_result_context(student, semester_no=semester_no)
    if error:
        return HttpResponse(error, status=404)

    response = HttpResponse(content_type='application/pdf')
    filename = f"Marksheet_{student.enrollment_no}_Sem{student.semester.number}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    from reportlab.lib.pagesizes import landscape
    p = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)
    y = height - 40

    sheet = result_context['sheet']

    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, y, "KADI SARVA VISHWAVIDYALAYA")
    y -= 18
    p.drawCentredString(width / 2, y, "STATEMENT OF MARKS / GRADE")

    y -= 18
    p.setFont("Helvetica", 12)
    p.drawCentredString(width / 2, y, f"Semester - {sheet.semester.number} Examination held in {sheet.exam_session}")

    y -= 20
    p.setFont("Helvetica", 12)
    p.drawString(40, y, f"Enrollment No: {student.enrollment_no}")
    y -= 12
    p.drawString(40, y, f"Student Name: {student.name}")
    y -= 12
    p.drawString(40, y, f"Branch: {student.branch.name}")

    y -= 18
    
    # Build main table with courses and remarks (grading scheme) on the right
    # Header row with nested structure
    main_table_data = []
    
    # Main header
    header_row1 = ["Course Code", "Subject Name", "Course Credit", "Grade", "Remarks"]
    main_table_data.append(header_row1)
    
    # Sub-header for remarks
    header_row2 = ["", "", "", "", ""]
    main_table_data.append(header_row2)
    
    # Course rows
    course_rows = []
    for row in result_context['rows']:
        course_rows.append([
            row['code'],
            row['name'],
            str(row['credit']),
            row['grade'],
            ""
        ])
    
    # Calculate how many rows we have for courses
    num_courses = len(course_rows)
    
    # Add course rows
    main_table_data.extend(course_rows)
    
    # Create the main table
    main_table = Table(main_table_data, colWidths=[1.2*inch, 3.2*inch, 1.0*inch, 0.8*inch, 2.5*inch])
    
    style_list = [
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('FONT', (0, 2), (-1, -1), 'Helvetica', 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (3, -1), 0.6, colors.black),
        ('GRID', (4, 0), (4, -1), 0.6, colors.black),
        ('BACKGROUND', (0, 0), (-1, 1), colors.whitesmoke),
        ('SPAN', (4, 0), (4, 1)),  # Merge Remarks header
        ('SPAN', (4, 2), (4, -1)),  # Merge all Remarks content cells
    ]
    
    main_table.setStyle(TableStyle(style_list))
    
    # Create grading scheme sub-table for remarks column
    scheme_data = [
        ["Grading\nScheme", "Percentage\nAccording to Grade", "Grade\nPoints"],
        ["A+", "90-100", "10"],
        ["A", "80-89", "9"],
        ["A-", "70-79", "8"],
        ["B+", "60-69", "7"],
        ["B", "50-59", "6"],
        ["B-", "40-49", "5"],
        ["F", "< 40", "0"],
    ]
    scheme_table = Table(scheme_data, colWidths=[0.8*inch, 1.0*inch, 0.6*inch])
    scheme_table.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 7),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 7),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
    ]))
    
    # Draw the main table
    main_table.wrapOn(p, width - 80, height)
    table_y = y - (len(main_table_data) * 16)
    main_table.drawOn(p, 40, table_y)
    
    # Calculate position for scheme table in the remarks column
    scheme_x = 40 + 1.2*inch + 3.2*inch + 1.0*inch + 0.8*inch + 0.1*inch
    scheme_y = table_y + (len(main_table_data) * 16) - 32 - (len(scheme_data) * 12)
    scheme_table.wrapOn(p, 2.4*inch, height)
    scheme_table.drawOn(p, scheme_x, scheme_y)
    
    y = table_y - 25

    # Summary tables side by side at bottom
    left_summary_data = [
        ["Credits", "Earned Grade\nPoints", "SPI"],
        [str(sheet.earned_credits or ''), str(sheet.earned_grade_points or ''), str(sheet.spi or '')],
    ]
    left_summary = Table(left_summary_data, colWidths=[1.0*inch, 1.2*inch, 0.8*inch])
    left_summary.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
    ]))
    left_summary.wrapOn(p, 3.0*inch, height)
    left_summary.drawOn(p, 40, y - 40)
    
    right_summary_data = [
        ["Earned\nCredits", "Earned Grade\nPoints", "CPI"],
        [str(sheet.total_credits or ''), str(sheet.total_grade_points or ''), str(sheet.cpi or '')],
    ]
    right_summary = Table(right_summary_data, colWidths=[1.0*inch, 1.2*inch, 0.8*inch])
    right_summary.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 0), 'Helvetica-Bold', 9),
        ('FONT', (0, 1), (-1, -1), 'Helvetica', 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
    ]))
    right_summary.wrapOn(p, 3.0*inch, height)
    right_summary.drawOn(p, 40 + 3.5*inch, y - 40)
    
    # Result at bottom
    y = y - 60
    p.setFont("Helvetica-Bold", 11)
    p.drawString(40, y, f"RESULT : {sheet.result_status}")
    
    if sheet.issued_date:
        p.setFont("Helvetica", 9)
        p.drawRightString(width - 40, y, f"Issued Date: {sheet.issued_date.strftime('%d %b %Y')}")

    p.setFont("Helvetica", 8)
    p.drawCentredString(width / 2, 20, "Generated by Exam Cell")

    p.save()
    return response
    p.drawCentredString(width / 2, 30, "Generated by Exam Cell")

    p.save()
    return response


# ============= ENROLLMENTS (TOGGLES) =============

def _subject_initials(name):
    if not name:
        return ''
    stop_words = {"and", "or", "the", "of", "for", "to", "a", "an", "in", "on", "with", "by", "from", "at", "as"}
    parts = [p for p in str(name).split() if p and p.lower() not in stop_words]
    if len(parts) == 1:
        return parts[0].upper()
    return ''.join(p[0].upper() for p in parts)

@login_required(login_url='login')
def manage_enrollments(request):
    """Manage enrollments for a branch+semester.
    Fixed subjects are auto-assigned; electives are selected per group."""
    if not request.user.is_staff:
        return redirect('dashboard')

    branches = Branch.objects.all().order_by('name')
    semesters = Semester.objects.all().order_by('number')

    branch_id = request.GET.get('branch') or request.POST.get('branch')
    semester_id = request.GET.get('semester') or request.POST.get('semester')
    selected_divisions = [d for d in request.GET.getlist('division') if d]
    if not selected_divisions:
        selected_divisions = [d for d in request.POST.getlist('division') if d]
    if not selected_divisions:
        single_division = request.GET.get('division') or request.POST.get('division')
        if single_division:
            selected_divisions = [single_division]
    filter_subject_id = request.GET.get('filter_subject') or request.POST.get('filter_subject')

    students = []
    subjects = []
    fixed_subjects = []
    elective_groups = []
    elective_selection = {}
    message = None
    upload_results = None
    upload_errors = []

    if branch_id and semester_id:
        qs = Student.objects.filter(branch_id=branch_id, semester_id=semester_id)
        subjects = list(Subject.objects.filter(branch_id=branch_id, semester_id=semester_id).order_by('code'))
        fixed_subjects = [s for s in subjects if not s.is_elective]

        elective_lookup = {}
        for subj in subjects:
            if not subj.is_elective:
                continue
            group_name = (subj.elective_group or '').strip() or 'Elective'
            elective_lookup.setdefault(group_name, []).append(subj)
        elective_groups = []
        for idx, (group_name, group_subjects) in enumerate(sorted(elective_lookup.items(), key=lambda item: item[0])):
            elective_groups.append({
                'key': str(idx),
                'name': group_name,
                'subjects': sorted(group_subjects, key=lambda s: s.code),
            })

        if selected_divisions:
            qs = qs.filter(division__in=selected_divisions)
        if filter_subject_id:
            qs = qs.filter(subjects_enrolled__subject_id=filter_subject_id).distinct()
        students = list(qs.order_by('division', 'enrollment_no'))

        # Auto-assign fixed subjects to all students in the selection.
        if students and fixed_subjects:
            fixed_pairs = {(stu.id, subj.id) for stu in students for subj in fixed_subjects}
            existing_fixed = set(
                StudentSubject.objects.filter(student__in=students, subject__in=fixed_subjects)
                .values_list('student_id', 'subject_id')
            )
            to_create = fixed_pairs - existing_fixed
            if to_create:
                StudentSubject.objects.bulk_create(
                    [StudentSubject(student_id=sid, subject_id=subid) for sid, subid in to_create],
                    ignore_conflicts=True,
                )

        if request.method == 'POST':
            action = request.POST.get('action')
            if action == 'save_electives':
                elective_subjects = [s for s in subjects if s.is_elective]
                elective_subject_ids = {s.id for s in elective_subjects}
                group_subject_ids = {
                    group['key']: {s.id for s in group['subjects']}
                    for group in elective_groups
                }

                selected = set()
                for stu in students:
                    for group in elective_groups:
                        field_name = f"elective_{stu.id}_{group['key']}"
                        value = (request.POST.get(field_name) or '').strip()
                        if not value:
                            continue
                        try:
                            subid = int(value)
                        except ValueError:
                            continue
                        if subid in group_subject_ids.get(group['key'], set()):
                            selected.add((stu.id, subid))

                existing = set(
                    StudentSubject.objects.filter(student__in=students, subject__in=elective_subjects)
                    .values_list('student_id', 'subject_id')
                )

                to_create = selected - existing
                if to_create:
                    StudentSubject.objects.bulk_create(
                        [StudentSubject(student_id=sid, subject_id=subid) for sid, subid in to_create],
                        ignore_conflicts=True,
                    )

                to_delete = existing - selected
                if to_delete:
                    StudentSubject.objects.filter(
                        student_id__in=[sid for sid, _ in to_delete],
                        subject_id__in=[subid for _, subid in to_delete if subid in elective_subject_ids],
                    ).delete()

                message = "Enrollments updated. Fixed subjects are assigned automatically."
            elif action == 'upload_electives':
                upload_file = request.FILES.get('file')
                if not upload_file:
                    upload_errors.append("Please select a file to upload.")
                else:
                    try:
                        if str(upload_file.name).lower().endswith('.csv'):
                            df = pd.read_csv(upload_file)
                        else:
                            df = pd.read_excel(upload_file)
                    except Exception as exc:
                        upload_errors.append(f"Failed to read file: {exc}")
                        df = None

                    if df is not None:
                        df.columns = [str(c).strip().lower().replace(' ', '_') for c in df.columns]
                        required = ['enrollment_no', 'elective_group', 'subject_code']
                        missing = [c for c in required if c not in df.columns]
                        if missing:
                            upload_errors.append(
                                f"Missing required columns: {', '.join(missing)}."
                            )
                        else:
                            def _norm_group(value):
                                return str(value or '').strip().upper()

                            elective_subjects = list(
                                Subject.objects.filter(
                                    branch_id=branch_id,
                                    semester_id=semester_id,
                                    is_elective=True,
                                )
                            )
                            subject_by_code = {s.code.strip().upper(): s for s in elective_subjects}
                            group_subjects = {}
                            for subj in elective_subjects:
                                group_key = _norm_group(subj.elective_group)
                                if group_key:
                                    group_subjects.setdefault(group_key, []).append(subj)

                            created = 0
                            updated = 0
                            skipped = 0

                            for idx, row in df.iterrows():
                                row_no = idx + 1
                                enroll = str(row.get('enrollment_no', '')).strip()
                                group_raw = row.get('elective_group', '')
                                code_raw = row.get('subject_code', '')

                                group_key = _norm_group(group_raw)
                                subject_code = str(code_raw or '').strip().upper()

                                if not enroll or not group_key or not subject_code:
                                    upload_errors.append(
                                        f"Row {row_no}: missing enrollment_no, elective_group, or subject_code."
                                    )
                                    continue

                                student = Student.objects.filter(
                                    enrollment_no__iexact=enroll,
                                    branch_id=branch_id,
                                    semester_id=semester_id,
                                ).first()
                                if not student:
                                    upload_errors.append(
                                        f"Row {row_no}: student not found in selected branch/semester ({enroll})."
                                    )
                                    continue

                                subject = subject_by_code.get(subject_code)
                                if not subject:
                                    upload_errors.append(
                                        f"Row {row_no}: elective subject code not found ({subject_code})."
                                    )
                                    continue

                                subject_group = _norm_group(subject.elective_group)
                                if not subject_group:
                                    upload_errors.append(
                                        f"Row {row_no}: subject has no elective group ({subject_code})."
                                    )
                                    continue
                                if subject_group != group_key:
                                    upload_errors.append(
                                        f"Row {row_no}: group mismatch for {subject_code} (expected {subject_group})."
                                    )
                                    continue

                                group_list = group_subjects.get(group_key, [])
                                if not group_list:
                                    upload_errors.append(
                                        f"Row {row_no}: elective group not found ({group_key})."
                                    )
                                    continue

                                existing_group = StudentSubject.objects.filter(
                                    student=student,
                                    subject__in=group_list,
                                )

                                if existing_group.filter(subject=subject).exists() and existing_group.count() == 1:
                                    skipped += 1
                                    continue

                                existing_group.exclude(subject=subject).delete()
                                _, was_created = StudentSubject.objects.get_or_create(
                                    student=student,
                                    subject=subject,
                                )
                                if was_created:
                                    created += 1
                                else:
                                    updated += 1

                            upload_results = {
                                'created': created,
                                'updated': updated,
                                'skipped': skipped,
                                'errors': len(upload_errors),
                                'total': len(df),
                            }

        # Refresh after changes: build map of student_id -> list of subject_ids
        pairs = StudentSubject.objects.filter(student__in=students, subject__in=subjects)
        pairs = pairs.values_list('student_id', 'subject_id')
        enroll_map = {}
        for sid, subid in pairs:
            enroll_map.setdefault(sid, []).append(subid)

        subject_group_key = {}
        for group in elective_groups:
            for subj in group['subjects']:
                subject_group_key[subj.id] = group['key']
        elective_selection = {}
        for sid, subid in pairs:
            group_key = subject_group_key.get(subid)
            if not group_key:
                continue
            elective_selection.setdefault(sid, {})[group_key] = subid
    else:
        enroll_map = {}

    if request.GET.get('download') == 'csv' and branch_id and semester_id:
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="enrollments.csv"'
        writer = csv.writer(response)
        header = ['Enrollment No', 'Name', 'Division']
        for s in subjects:
            initials = _subject_initials(s.name)
            header.append(f"{s.code} ({initials})" if initials else s.code)
        writer.writerow(header)
        for stu in students:
            enrolled = set(enroll_map.get(stu.id, []))
            row = [
                stu.enrollment_no,
                stu.name,
                stu.division,
            ]
            for subj in subjects:
                row.append('Yes' if subj.id in enrolled else 'No')
            writer.writerow(row)
        return response

    context = {
        'branches': branches,
        'semesters': semesters,
        'students': students,
        'subjects': subjects,
        'fixed_subjects': fixed_subjects,
        'elective_groups': elective_groups,
        'elective_selection': elective_selection,
        'branch_id': branch_id,
        'semester_id': semester_id,
        'selected_divisions': selected_divisions,
        'filter_subject_id': filter_subject_id,
        'enroll_map': enroll_map,
        'message': message,
        'upload_results': upload_results,
        'upload_errors': upload_errors,
    }
    return render(request, 'subjects/manage_enrollments.html', context)

def _build_attendance_rooms(semester_no):
    if not semester_no:
        return []

    seating = list(CESeating.objects.all().order_by('id'))
    students = list(
        Student.objects.filter(semester__number=semester_no)
        .select_related('branch', 'semester')
        .order_by('enrollment_no')
    )
    branch_name = students[0].branch.name if students else ''

    attendance_rooms = []
    idx = 0
    for room in seating:
        room_students = students[idx:idx + room.count]
        idx += room.count
        row_count = max(len(room_students), 32)
        rows = []
        for i in range(row_count):
            student = room_students[i] if i < len(room_students) else None
            rows.append({
                'sr': i + 1,
                'enrollment_no': student.enrollment_no if student else '',
                'name': student.name if student else '',
                'division': student.division if student else '',
            })
        attendance_rooms.append({
            'room_no': room.room_no,
            'branch_name': branch_name,
            'semester_no': semester_no,
            'rows': rows,
        })

    return attendance_rooms

@login_required(login_url='login')
def preview(request):
    """Preview exam seating"""
    seating = CESeating.objects.all().order_by('id')
    semester_raw = request.GET.get("semester", "")

    context = {
        "seating": seating,
        "semester": semester_raw,
        "month": request.GET.get("month", "NOVEMBER"),
        "year": request.GET.get("year", "2025"),
        "date_from": request.GET.get("date_from", ""),
        "date_to": request.GET.get("date_to", ""),
        "examination_name": request.GET.get("examination_name", ""),
        "note": request.GET.get("note", ""),
    }
    return render(request, "preview.html", context)

def _draw_attendance_sheet(p, meta, rows):
    width, height = A4
    y = height - 30
    right_block_x = width - 205
    right_block_end_x = width - 28

    def draw_right_field_line(y_pos, label, value=''):
        p.drawString(right_block_x, y_pos, label)
        label_width = p.stringWidth(label, "Helvetica", 9)
        line_start_x = right_block_x + label_width + 4
        if value:
            p.drawString(line_start_x, y_pos, value)
        else:
            p.line(line_start_x, y_pos - 2, right_block_end_x, y_pos - 2)

    def draw_inline_field(y_pos, x_start, x_end, label, value=''):
        p.drawString(x_start, y_pos, label)
        label_width = p.stringWidth(label, "Helvetica", 9)
        line_start_x = x_start + label_width + 3
        if value:
            p.drawString(line_start_x, y_pos, value)
        else:
            p.line(line_start_x, y_pos - 2, x_end, y_pos - 2)

    p.setFont("Helvetica-Bold", 12)
    p.drawCentredString(width / 2, y, meta['institute_line1'])
    y -= 14
    p.drawCentredString(width / 2, y, meta['institute_line2'])
    y -= 16
    p.setFont("Helvetica-Bold", 10)
    p.drawCentredString(width / 2, y, meta['exam_title'])
    y -= 14
    p.setFont("Helvetica-Bold", 11)
    p.drawCentredString(width / 2, y, meta['report_title'])

    y -= 14
    p.setFont("Helvetica", 9)
    branch_text = meta.get('branch') or '-'
    semester_text = meta.get('semester') or '-'
    p.drawString(40, y, f"Branch: {branch_text}")
    p.drawCentredString(width / 2, y, f"Semester: {semester_text}")
    block_value = meta.get('block_no') or ''
    room_value = (meta.get('room_no') or '') if meta.get('show_room_no', True) else ''
    right_mid_x = right_block_x + ((right_block_end_x - right_block_x) // 2)
    draw_inline_field(y, right_block_x, right_mid_x - 5, "Block:", block_value)
    draw_inline_field(y, right_mid_x + 5, right_block_end_x, "Room:", room_value)

    y -= 12
    subject_text = meta.get('subject') or "____________________________"
    p.drawString(40, y, f"Subject: {subject_text}")
    date_text = meta.get('date') or "______________"
    draw_right_field_line(y, "Date:", date_text if date_text != "______________" else "")

    y -= 15
    exam_time_text = (meta.get('exam_time') or '').strip()
    draw_right_field_line(y, "Exam Time:", exam_time_text)

    supervisor_name = meta.get('supervisor_name') or "____________________________"
    p.drawString(40, y, f"Full Name of Jr. Supervisor: {supervisor_name}")

    y -= 12
    p.drawString(40, y, "Supervisor Sign: ____________________________")

    y -= 14
    table_width = width - 80
    elective_count = int(meta.get('elective_count', 1) or 0)
    if elective_count <= 0:
        col_ratios = [0.06, 0.15, 0.40, 0.05, 0.075, 0.075, 0.075, 0.115]
    elif elective_count == 1:
        col_ratios = [0.06, 0.14, 0.36, 0.05, 0.06, 0.075, 0.075, 0.075, 0.105]
    else:
        col_ratios = [0.055, 0.13, 0.30, 0.05, 0.055, 0.055, 0.075, 0.075, 0.075, 0.13]
    col_widths = [table_width * r for r in col_ratios]

    base_cols = ["Sr. No", "Enrollment", "Name of Student", "Div"]
    if elective_count >= 1:
        base_cols.append("Elective-1" if elective_count > 1 else "Elective")
    if elective_count >= 2:
        base_cols.append("Elective-2")
    answersheet_start = len(base_cols)
    table_data = [
        base_cols + ["Answersheet", "", "", "Student Sign."],
        ["" for _ in base_cols] + ["Main", "Additional Supli", "", ""],
    ]

    for row in rows:
        data_row = [
            str(row['sr']),
            row['enrollment_no'],
            row['name'],
            row['division'],
        ]
        if elective_count >= 1:
            data_row.append(row.get('elective_1', ''))
        if elective_count >= 2:
            data_row.append(row.get('elective_2', ''))
        data_row.extend(["", "", "", ""])
        table_data.append(data_row)

    data_row_count = len(rows)
    header_row_heights = [14, 13]

    footer_min_top_y = 120
    available_table_height = max(120, y - footer_min_top_y)
    calculated_data_height = (available_table_height - sum(header_row_heights)) / max(data_row_count, 1)
    data_row_height = min(18, max(9, calculated_data_height))

    if data_row_height < 10:
        data_font_size = 6.5
        top_bottom_padding = 1
    elif data_row_height < 12:
        data_font_size = 7
        top_bottom_padding = 1
    elif data_row_height < 14:
        data_font_size = 7.5
        top_bottom_padding = 2
    else:
        data_font_size = 8
        top_bottom_padding = 3

    row_heights = header_row_heights + [data_row_height] * data_row_count
    t = Table(table_data, colWidths=col_widths, rowHeights=row_heights)
    header_spans = [
        ('SPAN', (answersheet_start, 0), (answersheet_start + 2, 0)),
        ('SPAN', (answersheet_start + 1, 1), (answersheet_start + 2, 1)),
    ]
    for col_idx in list(range(0, answersheet_start)) + [answersheet_start + 3]:
        header_spans.append(('SPAN', (col_idx, 0), (col_idx, 1)))

    t.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 1), 'Helvetica-Bold', 8),
        ('FONT', (0, 2), (-1, -1), 'Helvetica', data_font_size),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 2), (2, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.6, colors.black),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('TOPPADDING', (0, 0), (-1, -1), top_bottom_padding),
        ('BOTTOMPADDING', (0, 0), (-1, -1), top_bottom_padding),
    ] + header_spans))

    t.wrapOn(p, table_width, height)
    table_y = y - sum(row_heights)
    t.drawOn(p, 40, table_y)

    y = table_y - 18
    p.setFont("Helvetica", 9)
    p.drawString(40, y, "Enrollment No. of Absent Students:")
    y -= 14
    p.drawString(40, y, "1. ____________ 2. ____________ 3. ____________ 4. ____________ 5. ____________")
    y -= 12
    p.drawString(40, y, "6. ____________ 7. ____________ 8. ____________ 9. ____________ 10. ____________")

    y -= 16
    p.drawString(40, y, "Total Students: ____________")
    p.drawString(220, y, "Total Present: ____________")
    p.drawString(400, y, "Total Absent: ____________")

@login_required(login_url='login')
def download_pdf(request):
    """Download exam seating PDF"""
    # Allow faculty to download seating PDF
    
    semester = request.GET.get("semester", "")
    month = request.GET.get("month", "NOVEMBER")
    year = request.GET.get("year", "2025")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    note = request.GET.get("note", "")
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="CE_Seating.pdf"'
    
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    y = height - 40
    
    # HEADER
    p.setFont("Helvetica-Bold", 14)
    p.drawCentredString(width / 2, y, "KADI SARVA VISHWAVIDYALAYA")
    y -= 18
    p.drawCentredString(width / 2, y, "LDRP Institute of Technology and Research")
    
    # TITLE
    y -= 25
    box_height = 25
    p.setLineWidth(2)
    p.rect(40, y - box_height, width - 80, box_height)
    p.line(40 + (width - 80) / 3, y - box_height, 40 + (width - 80) / 3, y)
    p.line(40 + 2 * (width - 80) / 3, y - box_height, 40 + 2 * (width - 80) / 3, y)
    
    p.setFont("Helvetica-Bold", 13)
    p.drawString(50, y - 16, f"Semester:{semester}")
    p.drawCentredString(width / 2, y - 16, "Seating Arrangement")
    p.drawRightString(width - 50, y - 16, f"{month} {year}")
    
    y -= box_height + 5
    
    # DATE
    p.setFont("Helvetica-Bold", 11)
    p.setLineWidth(1.5)
    p.rect(40, y - 20, width - 80, 20)
    p.drawCentredString(width / 2, y - 13, f"Date: {date_from} to {date_to}")
    
    y -= 30
    
    # TABLE
    seating_data = CESeating.objects.all()
    table_data = [
        ['Branch', 'Block No', 'Room No', 'Seat Nos', '', 'No. of Students'],
        ['', '', '', 'From', 'To', '']
    ]
    
    for i, r in enumerate(seating_data, start=1):
        table_data.append(['', str(i), str(r.room_no), str(r.seat_from), str(r.seat_to), str(r.count)])
    
    if table_data[2:]:
        table_data[2][0] = 'CE'
    
    t = Table(table_data, colWidths=[1.2*inch, 0.8*inch, 0.9*inch, 1*inch, 1*inch, 1.1*inch])
    
    t.setStyle(TableStyle([
        ('FONT', (0, 0), (-1, 1), 'Helvetica-Bold', 11),
        ('ALIGN', (0, 0), (-1, 1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, 1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONT', (0, 2), (-1, -1), 'Helvetica', 10),
        ('ALIGN', (0, 2), (-1, -1), 'CENTER'),
    ]))

    table_width = 7.5 * inch
    t.wrapOn(p, table_width, height)
    t.drawOn(p, (width - table_width) / 2, y - (len(table_data) * 20))
    
    y = y - (len(table_data) * 20) - 20
    
    if note:
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, f"* {note}")
    
    # FOOTER
    p.setFont("Helvetica", 9)
    p.drawCentredString(width / 2, 30, "Prepared by Exam Section")
    
    p.save()
    return response


def download_seating_excel(request):
    """Download CE seating arrangement as Excel"""
    semester = request.GET.get("semester", "")
    month = request.GET.get("month", "NOVEMBER")
    year = request.GET.get("year", "2025")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    examination_name = request.GET.get("examination_name", "")
    note = request.GET.get("note", "")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
    except Exception:
        return HttpResponse('openpyxl is required for Excel export.', status=500)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "CE Seating"
    
    # Set column widths
    worksheet.column_dimensions['A'].width = 12
    worksheet.column_dimensions['B'].width = 12
    worksheet.column_dimensions['C'].width = 12
    worksheet.column_dimensions['D'].width = 12
    worksheet.column_dimensions['E'].width = 12
    worksheet.column_dimensions['F'].width = 18
    
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    
    # Header - Institution name
    worksheet.merge_cells('A1:F1')
    worksheet['A1'] = 'KADI SARVA VISHWAVIDYALAYA'
    worksheet['A1'].font = Font(bold=True, size=14)
    worksheet['A1'].alignment = Alignment(horizontal='center')
    
    worksheet.merge_cells('A2:F2')
    worksheet['A2'] = 'LDRP Institute of Technology and Research'
    worksheet['A2'].font = Font(bold=True, size=14)
    worksheet['A2'].alignment = Alignment(horizontal='center')
    
    if examination_name:
        worksheet.merge_cells('A3:F3')
        worksheet['A3'] = examination_name
        worksheet['A3'].font = Font(bold=True, size=11)
        worksheet['A3'].alignment = Alignment(horizontal='center')
        row = 4
    else:
        row = 3
    
    # Title row
    worksheet.merge_cells(f'A{row}:B{row}')
    worksheet[f'A{row}'] = f'Semester: {semester}'
    worksheet[f'A{row}'].font = Font(bold=True, size=13)
    worksheet[f'A{row}'].alignment = Alignment(horizontal='left')
    
    worksheet.merge_cells(f'C{row}:D{row}')
    worksheet[f'C{row}'] = 'Seating Arrangement'
    worksheet[f'C{row}'].font = Font(bold=True, size=13)
    worksheet[f'C{row}'].alignment = Alignment(horizontal='center')
    
    worksheet.merge_cells(f'E{row}:F{row}')
    worksheet[f'E{row}'] = f'{month} {year}'
    worksheet[f'E{row}'].font = Font(bold=True, size=13)
    worksheet[f'E{row}'].alignment = Alignment(horizontal='right')
    
    row += 1
    
    # Date row
    worksheet.merge_cells(f'A{row}:F{row}')
    worksheet[f'A{row}'] = f'Date: {date_from} to {date_to}'
    worksheet[f'A{row}'].font = Font(bold=True, size=11)
    worksheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    row += 2
    
    # Table headers - first row
    worksheet[f'A{row}'] = 'Branch'
    worksheet[f'B{row}'] = 'Block No'
    worksheet[f'C{row}'] = 'Room No'
    worksheet.merge_cells(f'D{row}:E{row}')
    worksheet[f'D{row}'] = 'Seat Nos'
    worksheet[f'F{row}'] = 'No. of Students'
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        worksheet[f'{col}{row}'].font = Font(bold=True, size=11)
        worksheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet[f'{col}{row}'].border = border
    
    row += 1
    
    # Table headers - second row
    worksheet[f'A{row}'] = ''
    worksheet[f'B{row}'] = ''
    worksheet[f'C{row}'] = ''
    worksheet[f'D{row}'] = 'From'
    worksheet[f'E{row}'] = 'To'
    worksheet[f'F{row}'] = ''
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        worksheet[f'{col}{row}'].font = Font(bold=True, size=11)
        worksheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        worksheet[f'{col}{row}'].border = border
    
    row += 1
    
    # Data rows
    seating_data = CESeating.objects.all().order_by('id')
    start_row = row
    
    for i, r in enumerate(seating_data, start=1):
        if i == 1:
            worksheet[f'A{row}'] = 'CE'
            worksheet.merge_cells(f'A{row}:A{start_row + len(seating_data) - 1}')
            worksheet[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        worksheet[f'B{row}'] = i
        worksheet[f'C{row}'] = r.room_no
        worksheet[f'D{row}'] = r.seat_from
        worksheet[f'E{row}'] = r.seat_to
        worksheet[f'F{row}'] = r.count
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            worksheet[f'{col}{row}'].border = border
            worksheet[f'{col}{row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Note
    if note:
        row += 1
        worksheet.merge_cells(f'A{row}:F{row}')
        worksheet[f'A{row}'] = f'* {note}'
        worksheet[f'A{row}'].font = Font(bold=True, size=10)
        worksheet[f'A{row}'].alignment = Alignment(horizontal='left')
    
    # Footer
    row += 2
    worksheet.merge_cells(f'A{row}:F{row}')
    worksheet[f'A{row}'] = 'Prepared by Exam Section'
    worksheet[f'A{row}'].font = Font(size=9)
    worksheet[f'A{row}'].alignment = Alignment(horizontal='center')
    
    # Save to response
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="CE_Seating_Arrangement.xlsx"'
    return response


@login_required(login_url='login')
def attendance_report(request):
    """Generate a single-page attendance report (optional student upload)."""
    if not request.user.is_staff and not hasattr(request.user, 'faculty'):
        return redirect('dashboard')

    attendance_branch_defaults = [
        ('Computer Engineering', 'CE'),
        ('Information Technology', 'IT'),
        ('Electronics and Communication Engineering', 'EC'),
        ('Mechanical Engineering', 'ME'),
        ('Automobile Engineering', 'AE'),
        ('Civil Engineering', ''),
        ('Electrical Engineering', 'EE'),
    ]

    branches = Branch.objects.all().order_by('name')
    existing_branch_names = {b.name.casefold() for b in branches}
    custom_branch_options = [
        {
            'name': name,
            'label': f"{name} ({code})" if code else name,
        }
        for name, code in attendance_branch_defaults
        if name.casefold() not in existing_branch_names
    ]
    branch_suggestions = []
    seen_branch_suggestions = set()
    for item in branches:
        branch_name = item.name
        key = branch_name.casefold()
        if key and key not in seen_branch_suggestions:
            seen_branch_suggestions.add(key)
            branch_suggestions.append(branch_name)
    for item in custom_branch_options:
        branch_name = item['name']
        key = branch_name.casefold()
        if key and key not in seen_branch_suggestions:
            seen_branch_suggestions.add(key)
            branch_suggestions.append(branch_name)
    semesters = Semester.objects.all().order_by('number')
    subjects = Subject.objects.select_related('branch', 'semester').order_by('semester__number', 'code')
    subject_suggestions = []
    seen_subject_suggestions = set()
    for subject_item in subjects:
        combined_label = f"{subject_item.code} - {subject_item.name}"
        for candidate in (combined_label, subject_item.name, subject_item.code):
            key = candidate.casefold()
            if key and key not in seen_subject_suggestions:
                seen_subject_suggestions.add(key)
                subject_suggestions.append(candidate)

    row_count = 35
    include_room_no = True
    selected_branch_id = ''
    selected_semester_id = ''
    selected_subject_id = ''
    subject_custom_input = ''
    selected_order_by = 'enrollment_no'
    exam_time_input = ''
    branch_custom_input = ''
    block_no_input = ''
    room_numbers_input = ''
    students_per_room_input = '35'
    students_per_room_list_input = ''
    can_download_summary = bool(request.session.get('attendance_summary_preview_done', False))

    message = None

    def _render_form(msg=None):
        return render(request, 'attendance/report_form.html', {
            'message': msg,
            'row_count': row_count,
            'include_room_no': include_room_no,
            'branches': branches,
            'semesters': semesters,
            'subjects': subjects,
            'selected_branch_id': selected_branch_id,
            'selected_semester_id': selected_semester_id,
            'selected_subject_id': selected_subject_id,
            'subject_custom_input': subject_custom_input,
            'selected_order_by': selected_order_by,
            'exam_time_input': exam_time_input,
            'branch_custom_input': branch_custom_input,
            'custom_branch_options': custom_branch_options,
            'branch_suggestions': branch_suggestions,
            'subject_suggestions': subject_suggestions,
            'block_no_input': block_no_input,
            'room_numbers_input': room_numbers_input,
            'students_per_room_input': students_per_room_input,
            'students_per_room_list_input': students_per_room_list_input,
            'can_download_summary': can_download_summary,
        })

    if request.method == 'POST':
        institute_line1 = (request.POST.get('institute_line1') or '').strip() or 'KADI SARVA VISHWAVIDYALAYA'
        institute_line2 = (request.POST.get('institute_line2') or '').strip() or 'LDRP Institute of Technology and Research Gandhinagar'
        exam_title = (request.POST.get('exam_title') or '').strip() or 'Mid Semester Examination'
        report_title = (request.POST.get('report_title') or '').strip() or 'ATTENDANCE REPORT'

        selected_branch_id = (request.POST.get('branch') or '').strip()
        branch_custom_input = (request.POST.get('branch_custom') or '').strip()
        selected_semester_id = (request.POST.get('semester') or '').strip()
        selected_subject_id = (request.POST.get('subject') or '').strip()
        subject_custom_input = (request.POST.get('subject_custom') or '').strip()
        block_no_input = (request.POST.get('block_no') or '').strip()
        room_numbers_input = (request.POST.get('room_numbers') or '').strip()
        students_per_room_input = (request.POST.get('students_per_room') or '').strip()
        students_per_room_list_input = (request.POST.get('students_per_room_list') or '').strip()
        date_text = (request.POST.get('date_text') or '').strip()
        exam_time_input = (request.POST.get('exam_time') or '').strip()
        supervisor_name = (request.POST.get('supervisor_name') or '').strip()
        selected_order_by = (request.POST.get('order_by') or 'enrollment_no').strip()
        if selected_order_by not in ('uploaded', 'enrollment_no', 'exam_no', 'name'):
            selected_order_by = 'enrollment_no'
        output_format = (request.POST.get('output_format') or 'pdf').strip().lower()
        if output_format not in ('pdf', 'excel', 'summary_preview', 'summary_download'):
            output_format = 'pdf'
        include_room_no = request.POST.get('include_room_no') == 'on'
        students_file = request.FILES.get('students_file')

        try:
            row_count = int(request.POST.get('row_count') or 35)
        except ValueError:
            row_count = 35
        row_count = max(1, min(40, row_count))

        branch_obj = None
        sem_obj = None
        subject_obj = None
        selected_branch_label = ''

        if selected_branch_id:
            if selected_branch_id.startswith('custom:'):
                selected_branch_label = selected_branch_id.split(':', 1)[1].strip()
            else:
                branch_obj = Branch.objects.filter(id=selected_branch_id).first()
                if not branch_obj:
                    message = 'Invalid branch selected.'
                else:
                    selected_branch_label = branch_obj.name

        if not message and branch_custom_input:
            selected_branch_label = branch_custom_input
            matched_branch = Branch.objects.filter(name__iexact=branch_custom_input).first()
            if not matched_branch:
                matched_branch = Branch.objects.filter(code__iexact=branch_custom_input).first()
            if matched_branch:
                branch_obj = matched_branch
                selected_branch_id = str(matched_branch.id)

        if not message and selected_semester_id:
            sem_obj = Semester.objects.filter(id=selected_semester_id).first()
            if not sem_obj:
                message = 'Invalid semester selected.'

        if not message and selected_subject_id:
            subject_obj = Subject.objects.filter(id=selected_subject_id).first()
            if not subject_obj:
                message = 'Invalid subject selected.'

        block_tokens = [token.strip() for token in re.split(r'[\n,]+', block_no_input) if token.strip()]

        def _block_for_page(page_index):
            if not block_tokens:
                return block_no_input
            if page_index < len(block_tokens):
                return block_tokens[page_index]
            return block_tokens[-1]

        students_per_room = 0
        room_count_overrides = []
        room_labels = []
        if not message:
            if not room_numbers_input:
                message = 'Please enter room numbers.'
            if not message:
                try:
                    students_per_room = int(students_per_room_input)
                except Exception:
                    students_per_room = 0
                if students_per_room <= 0:
                    message = 'Students per room must be a positive number.'
            if not message:
                normalized_rooms = room_numbers_input.replace('\n', ',')
                try:
                    room_labels = parse_rooms(normalized_rooms)
                except Exception:
                    room_labels = []
                if not room_labels:
                    message = 'No valid room numbers found. Example: 101,102,103 or 301-305.'
            if not message and students_per_room_list_input:
                raw_tokens = [token.strip() for token in re.split(r'[\n,]+', students_per_room_list_input) if token.strip()]
                if not raw_tokens:
                    message = 'Room-wise students count is invalid. Example: 30,35,32.'
                else:
                    try:
                        room_count_overrides = [int(token) for token in raw_tokens]
                    except Exception:
                        room_count_overrides = []
                    if not room_count_overrides:
                        message = 'Room-wise students count is invalid. Example: 30,35,32.'
                    elif any(count <= 0 for count in room_count_overrides):
                        message = 'Each room-wise student count must be a positive number.'
                    elif len(room_count_overrides) > len(room_labels):
                        message = 'Room-wise count entries cannot exceed number of rooms.'

        rows = []
        use_uploaded_students = bool(students_file)

        if output_format != 'summary_download':
            request.session['attendance_summary_preview_done'] = False
            request.session.modified = True

        if use_uploaded_students:
            if selected_order_by == 'enrollment_no':
                selected_order_by = 'uploaded'
            rows, upload_error = _parse_attendance_students_file(students_file)
            if upload_error:
                message = upload_error
            else:
                if selected_order_by == 'name':
                    rows.sort(key=lambda r: (_clean_cell(r.get('name')).lower(), _clean_cell(r.get('enrollment_no')).lower()))
                elif selected_order_by == 'enrollment_no':
                    rows.sort(key=lambda r: _clean_cell(r.get('enrollment_no')).lower())
                elif selected_order_by == 'exam_no':
                    def exam_key(record):
                        exam_value = _clean_cell(record.get('exam_no'))
                        try:
                            return (0, int(exam_value))
                        except Exception:
                            return (1, exam_value.lower())
                    rows.sort(key=lambda r: (exam_key(r), _clean_cell(r.get('enrollment_no')).lower()))
                # selected_order_by == 'uploaded' keeps source row order from uploaded file
        else:
            students = Student.objects.all()

            if not message and branch_obj:
                students = students.filter(branch=branch_obj)

            if not message and sem_obj:
                students = students.filter(semester=sem_obj)

            if not message:
                order_field = 'enrollment_no'
                if selected_order_by == 'exam_no':
                    order_field = 'exam_no'
                elif selected_order_by == 'name':
                    order_field = 'name'

                students = list(students.order_by(order_field, 'enrollment_no'))
                student_ids = [s.id for s in students]
                elective_map = {}
                if student_ids:
                    elective_qs = (
                        StudentSubject.objects
                        .filter(student_id__in=student_ids, subject__is_elective=True)
                        .select_related('subject')
                        .order_by('student_id', 'subject__code')
                    )
                    for ss in elective_qs:
                        label = ss.subject.code
                        elective_map.setdefault(ss.student_id, []).append(label)

                for s in students:
                    electives = elective_map.get(s.id, [])
                    rows.append({
                        'sr': 0,
                        'enrollment_no': s.enrollment_no,
                        'name': s.name,
                        'division': s.division,
                        'elective_1': electives[0] if len(electives) > 0 else '',
                        'elective_2': electives[1] if len(electives) > 1 else '',
                        'room_no': s.room_no or '',
                        'exam_no': s.exam_no or '',
                    })
                if not rows:
                    message = "No students found for the selected filters."

        if message:
            return _render_form(message)

        branch = selected_branch_label or (branch_obj.name if branch_obj else '')
        semester = str(sem_obj.number) if sem_obj else ''
        if sem_obj is None:
            elective_count = 1
        elif sem_obj.number <= 4:
            elective_count = 0
        else:
            elective_count = 1
        if subject_custom_input:
            subject = subject_custom_input
        elif subject_obj:
            subject = f"{subject_obj.code} - {subject_obj.name}"
        else:
            subject = ''

        meta_base = {
            'institute_line1': institute_line1,
            'institute_line2': institute_line2,
            'exam_title': exam_title,
            'report_title': report_title,
            'branch': branch,
            'semester': semester,
            'subject': subject,
            'date': date_text,
            'exam_time': exam_time_input,
            'supervisor_name': supervisor_name,
            'block_no': block_no_input,
            'show_room_no': include_room_no,
            'elective_count': elective_count,
        }
        include_elective = elective_count >= 1

        room_specs = []
        for idx, room_label in enumerate(room_labels):
            room_count = students_per_room
            if room_count_overrides:
                if idx < len(room_count_overrides):
                    room_count = room_count_overrides[idx]
            room_specs.append({
                'room_no': str(room_label),
                'count': room_count,
            })

        pages = []
        pointer = 0
        total_capacity = sum(spec['count'] for spec in room_specs)
        if len(rows) > total_capacity:
            return _render_form(f"Student list has {len(rows)} rows, but seating capacity is {total_capacity}.")

        for page_index, spec in enumerate(room_specs):
            count = spec['count']
            if count > row_count:
                return _render_form(f"Room {spec['room_no']} has {count} students. Max {row_count} rows.")
            room_rows = rows[pointer:pointer + count]
            if not room_rows:
                break
            pointer += count
            range_from = _clean_cell(room_rows[0].get('enrollment_no'))
            range_to = _clean_cell(room_rows[-1].get('enrollment_no'))
            pages.append((spec['room_no'], room_rows, range_from, range_to, _block_for_page(page_index)))

        if output_format in ('summary_preview', 'summary_download'):
            def enrollment_group_key(enrollment_no):
                token = re.sub(r'[^A-Za-z0-9]', '', _clean_cell(enrollment_no)).upper()
                batch_match = re.match(r'^(\d+)', token)
                batch = batch_match.group(1) if batch_match else ''
                is_d2d = bool(re.match(r'^\d{2,3}S', token))
                return f"{batch}|{'D2D' if is_d2d else 'REG'}"

            branch_label = (branch_obj.code if branch_obj else '') or (branch or '-')
            table_rows = []
            total_students = 0
            first_branch_row = True

            for room_label, room_rows, _, _, block_value in pages:
                grouped = []
                for student_row in room_rows:
                    enrollment_no = _clean_cell(student_row.get('enrollment_no'))
                    if not enrollment_no:
                        continue
                    key = enrollment_group_key(enrollment_no)
                    if not grouped or grouped[-1]['key'] != key:
                        grouped.append({
                            'key': key,
                            'from': enrollment_no,
                            'to': enrollment_no,
                            'count': 1,
                        })
                    else:
                        grouped[-1]['to'] = enrollment_no
                        grouped[-1]['count'] += 1

                if not grouped:
                    grouped = [{'key': 'NA', 'from': '-', 'to': '-', 'count': 0}]

                for group_index, group_row in enumerate(grouped):
                    table_rows.append({
                        'branch': branch_label,
                        'show_branch': first_branch_row,
                        'block_no': str(block_value) if group_index == 0 else '',
                        'room_no': str(room_label) if group_index == 0 else '',
                        'seat_from': group_row['from'],
                        'seat_to': group_row['to'],
                        'count': str(group_row['count']),
                        'is_total': False,
                    })
                    total_students += group_row['count']
                    if first_branch_row:
                        first_branch_row = False

            # Calculate branch rowspan (total data rows excluding total row)
            branch_rowspan = len(table_rows)
            if table_rows:
                table_rows[0]['branch_rowspan'] = branch_rowspan

            table_rows.append({
                'branch': '',
                'show_branch': False,
                'block_no': '',
                'room_no': '',
                'seat_from': '',
                'seat_to': 'Total',
                'count': str(total_students),
                'is_total': True,
            })

            if output_format == 'summary_preview':
                request.session['attendance_summary_preview_done'] = True
                request.session['attendance_summary_payload'] = {
                    'institute_line1': institute_line1,
                    'institute_line2': institute_line2,
                    'exam_title': exam_title,
                    'semester': semester,
                    'table_rows': table_rows,
                }
                request.session.modified = True
                return render(request, 'attendance/seating_summary_preview.html', {
                    'institute_line1': institute_line1,
                    'institute_line2': institute_line2,
                    'exam_title': exam_title,
                    'semester': semester,
                    'branch': branch,
                    'table_rows': table_rows,
                })

            if not can_download_summary:
                return _render_form('Please preview seating summary first, then click download.')
            response = _build_attendance_summary_pdf_response(
                institute_line1=institute_line1,
                institute_line2=institute_line2,
                exam_title=exam_title,
                semester=semester,
                table_rows=table_rows,
            )
            request.session['attendance_summary_preview_done'] = False
            request.session.modified = True
            return response

        if output_format == 'excel':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, Border, Side
            except Exception:
                return _render_form('openpyxl is not installed. Please install it to export Excel.')

            workbook = Workbook()
            workbook.remove(workbook.active)
            thin = Side(style='thin', color='000000')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for sheet_idx, (room_label, room_rows, range_from, range_to, block_value) in enumerate(pages, start=1):
                sheet_name = f"Room {room_label}"
                worksheet = workbook.create_sheet(title=sheet_name[:31] or f"Sheet {sheet_idx}")

                if include_elective:
                    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']
                    worksheet.column_dimensions['A'].width = 8
                    worksheet.column_dimensions['B'].width = 18
                    worksheet.column_dimensions['C'].width = 32
                    worksheet.column_dimensions['D'].width = 14
                    worksheet.column_dimensions['E'].width = 8
                    worksheet.column_dimensions['F'].width = 12
                    worksheet.column_dimensions['G'].width = 14
                    worksheet.column_dimensions['H'].width = 14
                    worksheet.column_dimensions['I'].width = 16
                    last_col = 'I'
                else:
                    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
                    worksheet.column_dimensions['A'].width = 8
                    worksheet.column_dimensions['B'].width = 18
                    worksheet.column_dimensions['C'].width = 32
                    worksheet.column_dimensions['D'].width = 8
                    worksheet.column_dimensions['E'].width = 12
                    worksheet.column_dimensions['F'].width = 14
                    worksheet.column_dimensions['G'].width = 14
                    worksheet.column_dimensions['H'].width = 16
                    last_col = 'H'

                worksheet.merge_cells(f'A1:{last_col}1')
                worksheet['A1'] = institute_line1
                worksheet['A1'].font = Font(bold=True, size=12)
                worksheet['A1'].alignment = Alignment(horizontal='center')

                worksheet.merge_cells(f'A2:{last_col}2')
                worksheet['A2'] = institute_line2
                worksheet['A2'].font = Font(bold=True, size=11)
                worksheet['A2'].alignment = Alignment(horizontal='center')

                worksheet.merge_cells(f'A3:{last_col}3')
                worksheet['A3'] = exam_title
                worksheet['A3'].font = Font(bold=True, size=10)
                worksheet['A3'].alignment = Alignment(horizontal='center')

                worksheet.merge_cells(f'A4:{last_col}4')
                worksheet['A4'] = report_title
                worksheet['A4'].font = Font(bold=True, size=11)
                worksheet['A4'].alignment = Alignment(horizontal='center')

                worksheet['A6'] = f"Branch: {branch or '-'}"
                worksheet['D6'] = f"Semester: {semester or '-'}"
                worksheet['G6'] = (
                    f"Block No.: {block_value or '____________'}   "
                    f"Room No.: {room_label if include_room_no else '____________'}"
                )
                worksheet.merge_cells('A8:D8')
                worksheet['A8'] = f"Subject: {subject or '____________________________'}"
                if include_elective:
                    worksheet.merge_cells('G7:I7')
                    worksheet['G7'] = f"Date: {date_text or '______________'}"
                    worksheet.merge_cells('G9:I9')
                    worksheet['G9'] = f"Exam Time: {exam_time_input or '______________'}"
                    worksheet.merge_cells('A9:F9')
                    worksheet['A9'] = f"Full Name of Jr. Supervisor: {supervisor_name or '____________________________'}"
                    worksheet.merge_cells('A10:F10')
                    worksheet['A10'] = "Supervisor Sign: ____________________________"

                    worksheet['A12'] = 'Sr. No'
                    worksheet['B12'] = 'Enrollment'
                    worksheet['C12'] = 'Name of Student'
                    worksheet['D12'] = 'Div'
                    worksheet['E12'] = 'Elective'
                    worksheet['F12'] = 'Answersheet'
                    worksheet.merge_cells('F12:H12')
                    worksheet['I12'] = 'Student Sign.'

                    worksheet['F13'] = 'Main'
                    worksheet['G13'] = 'Additional Supli'
                    worksheet.merge_cells('G13:H13')
                else:
                    worksheet.merge_cells('G7:H7')
                    worksheet['G7'] = f"Date: {date_text or '______________'}"
                    worksheet.merge_cells('G9:H9')
                    worksheet['G9'] = f"Exam Time: {exam_time_input or '______________'}"
                    worksheet.merge_cells('A9:E9')
                    worksheet['A9'] = f"Full Name of Jr. Supervisor: {supervisor_name or '____________________________'}"
                    worksheet.merge_cells('A10:E10')
                    worksheet['A10'] = "Supervisor Sign: ____________________________"

                    worksheet['A12'] = 'Sr. No'
                    worksheet['B12'] = 'Enrollment'
                    worksheet['C12'] = 'Name of Student'
                    worksheet['D12'] = 'Div'
                    worksheet['E12'] = 'Answersheet'
                    worksheet.merge_cells('E12:G12')
                    worksheet['H12'] = 'Student Sign.'

                    worksheet['E13'] = 'Main'
                    worksheet['F13'] = 'Additional Supli'
                    worksheet.merge_cells('F13:G13')

                for col in columns:
                    worksheet[f'{col}12'].font = Font(bold=True, size=9)
                    worksheet[f'{col}12'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    worksheet[f'{col}13'].font = Font(bold=True, size=9)
                    worksheet[f'{col}13'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                filled = []
                for i, r in enumerate(list(room_rows)[:row_count], start=1):
                    item = {
                        'sr': i,
                        'enrollment_no': r.get('enrollment_no', ''),
                        'name': r.get('name', ''),
                        'division': r.get('division', ''),
                    }
                    if include_elective:
                        item['elective'] = r.get('elective_1', '')
                    filled.append(item)
                while len(filled) < row_count:
                    item = {
                        'sr': len(filled) + 1,
                        'enrollment_no': '',
                        'name': '',
                        'division': '',
                    }
                    if include_elective:
                        item['elective'] = ''
                    filled.append(item)

                start_row = 14
                for idx, item in enumerate(filled):
                    row_no = start_row + idx
                    worksheet[f'A{row_no}'] = item['sr']
                    worksheet[f'B{row_no}'] = item['enrollment_no']
                    worksheet[f'C{row_no}'] = item['name']
                    if include_elective:
                        worksheet[f'D{row_no}'] = item['division']
                        worksheet[f'E{row_no}'] = item['elective']
                        worksheet[f'F{row_no}'] = ''
                        worksheet[f'G{row_no}'] = ''
                        worksheet[f'H{row_no}'] = ''
                        worksheet[f'I{row_no}'] = ''
                    else:
                        worksheet[f'D{row_no}'] = item['division']
                        worksheet[f'E{row_no}'] = ''
                        worksheet[f'F{row_no}'] = ''
                        worksheet[f'G{row_no}'] = ''
                        worksheet[f'H{row_no}'] = ''

                end_row = start_row + row_count - 1
                for row_no in range(12, end_row + 1):
                    for col in columns:
                        cell = worksheet[f'{col}{row_no}']
                        cell.border = border
                        if row_no >= start_row:
                            cell.alignment = Alignment(horizontal='center', vertical='center')

                for row_no in range(start_row, end_row + 1):
                    worksheet[f'C{row_no}'].alignment = Alignment(horizontal='left', vertical='center')
                    worksheet[f'E{row_no}'].alignment = Alignment(horizontal='left', vertical='center')

                footer_row = end_row + 2
                worksheet.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=15)
                worksheet[f'A{footer_row}'] = 'Enrollment No. of Absent Students: 1. ________________ 2. ________________ 3. ________________ 4. ________________ 5. ________________'
                worksheet[f'A{footer_row + 1}'] = '6. ________________ 7. ________________ 8. ________________ 9. ________________ 10. ________________'
                worksheet[f'A{footer_row + 3}'] = 'Total Students: ____________'
                worksheet[f'D{footer_row + 3}'] = 'Total Present: ____________'
                worksheet[f'G{footer_row + 3}'] = 'Total Absent: ____________'

            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="attendance_report_format.xlsx"'
            return response

        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="attendance_report_bulk.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        for idx, (room_label, room_rows, range_from, range_to, block_value) in enumerate(pages):
            room_rows = list(room_rows)
            filled = []
            for i, r in enumerate(room_rows[:row_count], start=1):
                filled.append({
                    'sr': i,
                    'enrollment_no': r.get('enrollment_no', ''),
                    'name': r.get('name', ''),
                    'division': r.get('division', ''),
                    'elective_1': r.get('elective_1', ''),
                    'elective_2': r.get('elective_2', ''),
                })
            while len(filled) < row_count:
                filled.append({
                    'sr': len(filled) + 1,
                    'enrollment_no': '',
                    'name': '',
                    'division': '',
                    'elective_1': '',
                    'elective_2': '',
                })

            meta = dict(meta_base)
            meta['room_no'] = room_label
            meta['block_no'] = block_value
            meta['range_from'] = range_from
            meta['range_to'] = range_to
            _draw_attendance_sheet(p, meta, filled)
            if idx < len(pages) - 1:
                p.showPage()
        p.save()
        return response

    return _render_form()


@login_required
def attendance_summary_download(request):
    payload = request.session.get('attendance_summary_payload') or {}
    can_download = bool(request.session.get('attendance_summary_preview_done', False))
    if not can_download or not payload:
        return redirect('attendance_report')

    response = _build_attendance_summary_pdf_response(
        institute_line1=payload.get('institute_line1') or 'KADI SARVA VISHWAVIDYALAYA',
        institute_line2=payload.get('institute_line2') or 'LDRP Institute of Technology and Research Gandhinagar',
        exam_title=payload.get('exam_title') or 'Mid Semester Examination',
        semester=payload.get('semester') or '',
        table_rows=payload.get('table_rows') or [],
    )
    request.session['attendance_summary_preview_done'] = False
    request.session.modified = True
    return response
